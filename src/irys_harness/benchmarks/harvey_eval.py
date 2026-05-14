from __future__ import annotations

import json
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from irys_harness.artifacts import safe_text

EVAL_XLSX_CELL_CHAR_LIMIT = 500


@dataclass(frozen=True)
class HarveyEvalPackage:
    run_id: str
    task_id: str
    run_dir: str
    output_dir: str
    copied_files: list[str]
    validation: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "run_id": self.run_id,
            "task_id": self.task_id,
            "run_dir": self.run_dir,
            "output_dir": self.output_dir,
            "copied_files": self.copied_files,
            "validation": self.validation,
        }


def prepare_harvey_eval_package(
    trace: dict[str, Any],
    *,
    harvey_root: str | Path,
    run_id: str | None = None,
) -> HarveyEvalPackage:
    if trace.get("benchmark") != "harvey_lab_sample":
        raise ValueError("Trace is not a Harvey LAB trace")
    task_id = str(trace["task_id"])
    resolved_run_id = run_id or f"irys/{trace['run_id']}"
    run_dir = Path(harvey_root) / "results" / resolved_run_id
    output_dir = run_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    copied: list[str] = []
    validation: list[dict[str, Any]] = []
    for artifact in trace.get("artifacts", []):
        source = Path(artifact["path"])
        if not source.exists():
            continue
        target = output_dir / source.name
        validation.append(copy_eval_artifact(source, target))
        copied.append(str(target))

    return HarveyEvalPackage(
        run_id=resolved_run_id,
        task_id=task_id,
        run_dir=str(run_dir),
        output_dir=str(output_dir),
        copied_files=copied,
        validation=validation,
    )


def copy_eval_artifact(source: Path, target: Path) -> dict[str, Any]:
    target.parent.mkdir(parents=True, exist_ok=True)
    suffix = source.suffix.lower()
    record: dict[str, Any] = {
        "source": str(source),
        "target": str(target),
        "extension": suffix,
        "source_bytes": source.stat().st_size if source.exists() else 0,
        "mode": "raw_copy",
        "issues": [],
    }
    if suffix in {".xlsx", ".xlsm"}:
        try:
            sanitize_xlsx_for_eval(source, target)
            record["mode"] = "sanitized_xlsx"
            record["target_bytes"] = target.stat().st_size if target.exists() else 0
            record["cell_char_limit"] = EVAL_XLSX_CELL_CHAR_LIMIT
            return record
        except Exception as exc:  # noqa: BLE001 - fallback keeps scoring restartable.
            record["issues"].append(f"{type(exc).__name__}: {exc}")
    shutil.copy2(source, target)
    record["target_bytes"] = target.stat().st_size if target.exists() else 0
    return record


def sanitize_xlsx_for_eval(source: Path, target: Path) -> None:
    workbook = load_workbook(source)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = safe_text(cell.value, limit=EVAL_XLSX_CELL_CHAR_LIMIT)
        workbook.save(target)
    finally:
        workbook.close()


def evaluate_prepared_harvey_run(
    *,
    harvey_root: str | Path,
    run_id: str,
    task_id: str,
    judge_model: str,
    parallel: int = 24,
) -> dict[str, Any]:
    root = Path(harvey_root).resolve()
    sys.path.insert(0, str(root))
    try:
        from evaluation.judge import Judge  # type: ignore
        from evaluation.run_eval import evaluate_run  # type: ignore

        judge = Judge(model=judge_model)
        try:
            return evaluate_run(run_id=run_id, task=task_id, judge=judge, parallel=parallel)
        except UnicodeDecodeError:
            return evaluate_harvey_run_utf8(
                harvey_root=root,
                run_id=run_id,
                task_id=task_id,
                judge=judge,
                parallel=parallel,
            )
    finally:
        try:
            sys.path.remove(str(root))
        except ValueError:
            pass


def evaluate_harvey_run_utf8(
    *,
    harvey_root: Path,
    run_id: str,
    task_id: str,
    judge: Any,
    parallel: int,
) -> dict[str, Any]:
    from evaluation.run_eval import validate_task_config  # type: ignore
    from evaluation.scoring import score_rubric  # type: ignore

    task_dir = harvey_root / "tasks" / Path(*task_id.split("/"))
    run_dir = harvey_root / "results" / run_id
    config_path = task_dir / "task.json"
    if not config_path.exists():
        raise FileNotFoundError(f"task.json not found: {config_path}")
    config = json.loads(config_path.read_text(encoding="utf-8", errors="replace"))
    validate_task_config(config=config, task_path=config_path)
    if not run_dir.exists():
        raise FileNotFoundError(f"run directory not found: {run_dir}")

    result = score_rubric(
        criteria=config["criteria"],
        run_dir=run_dir,
        judge=judge,
        task_desc=config["title"],
        parallel=parallel,
    )
    n_criteria = len(result.criteria_results)
    n_passed = sum(1 for c in result.criteria_results if c["verdict"] == "pass")
    all_pass = n_criteria > 0 and n_passed == n_criteria
    scores = {
        "score": result.score,
        "max_score": result.max_score,
        "summary": (
            f"{n_passed}/{n_criteria} criteria passed."
            + ("  ALL-PASS." if all_pass else f"  Missed {n_criteria - n_passed} - task FAIL.")
        ),
        "all_pass": all_pass,
        "n_criteria": n_criteria,
        "n_passed": n_passed,
        "criteria_results": result.criteria_results,
        "run_id": run_id,
        "task": task_id,
        "judge_model": judge.model,
        "scored_at": datetime.now(timezone.utc).isoformat(),
        "irys_eval_notes": ["Loaded Harvey task.json with explicit UTF-8 fallback."],
    }
    (run_dir / "scores.json").write_text(json.dumps(scores, indent=2), encoding="utf-8")
    return scores
