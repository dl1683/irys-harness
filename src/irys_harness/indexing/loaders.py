from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from markitdown import MarkItDown
from openpyxl import load_workbook


PDF_EXTRACTION_GLYPH_MAP = str.maketrans(
    {
        "\ue534": "0",
        "\ue536": "1",
        "\ue537": "2",
        "\ue538": "3",
        "\ue539": "4",
        "\ue53a": "5",
        "\ue53b": "6",
        "\ue53c": "7",
        "\ue53d": "8",
        "\ue53e": "9",
        "\ue389": "ff",
        "\ue38c": "fi",
        "\ue38d": "fl",
    }
)


@dataclass(frozen=True)
class DocumentRecord:
    doc_id: str
    path: str
    filename: str
    extension: str
    bytes: int | None
    source_posture: str
    text_chars: int = 0
    load_error: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "doc_id": self.doc_id,
            "path": self.path,
            "filename": self.filename,
            "extension": self.extension,
            "bytes": self.bytes,
            "source_posture": self.source_posture,
            "text_chars": self.text_chars,
            "load_error": self.load_error,
        }


@dataclass(frozen=True)
class ChunkRecord:
    chunk_id: str
    doc_id: str
    index: int
    text: str
    prev_chunk: str | None = None
    next_chunk: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "chunk_id": self.chunk_id,
            "doc_id": self.doc_id,
            "index": self.index,
            "text": self.text,
            "prev_chunk": self.prev_chunk,
            "next_chunk": self.next_chunk,
        }


def load_documents(
    paths: list[str],
    *,
    chunk_size_chars: int = 4000,
    chunk_overlap_chars: int = 400,
    max_chars_per_doc: int | None = 200_000,
    source_posture: str = "benchmark_provided_context",
    progress_callback: Callable[[dict[str, Any]], None] | None = None,
    should_cancel: Callable[[], bool] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    documents: list[dict[str, Any]] = []
    chunks: list[dict[str, Any]] = []
    converter = MarkItDown()
    total = len(paths)
    for index, raw_path in enumerate(paths, 1):
        if should_cancel and should_cancel():
            raise RuntimeError("run canceled")
        path = Path(raw_path)
        emit_progress(
            progress_callback,
            "READ",
            "Reading document",
            summary=f"Reading document {index} of {total}: {path.name}",
            current=index,
            total=total,
            filename=path.name,
            path=str(path),
            bytes=path.stat().st_size if path.exists() else None,
            steer_hint="If this is not the right source, stop the run and remove or replace it before rerunning.",
        )
        document, text = load_document(
            path,
            doc_id=f"doc_{index:04d}",
            converter=converter,
            max_chars=max_chars_per_doc,
            source_posture=source_posture,
        )
        documents.append(document.to_dict())
        doc_chunks = chunk_text(
            text,
            doc_id=document.doc_id,
            chunk_size_chars=chunk_size_chars,
            chunk_overlap_chars=chunk_overlap_chars,
        )
        chunks.extend(chunk.to_dict() for chunk in doc_chunks)
        emit_progress(
            progress_callback,
            "READ",
            "Finished document",
            summary=f"Finished {path.name}",
            current=index,
            total=total,
            filename=path.name,
            doc_id=document.doc_id,
            text_chars=document.text_chars,
            chunk_count=len(doc_chunks),
            load_error=document.load_error,
        )
    return documents, chunks


def emit_progress(
    progress_callback: Callable[[dict[str, Any]], None] | None,
    label: str,
    message: str,
    **fields: Any,
) -> None:
    if progress_callback is None:
        return
    progress_callback({"label": label, "message": message, "fields": fields})


def load_document(
    path: Path,
    *,
    doc_id: str,
    converter: MarkItDown | None = None,
    max_chars: int | None = 200_000,
    source_posture: str = "benchmark_provided_context",
) -> tuple[DocumentRecord, str]:
    try:
        text = normalize_extracted_text(convert_to_text(path, converter=converter))
        if max_chars is not None:
            text = text[:max_chars]
        return (
            DocumentRecord(
                doc_id=doc_id,
                path=str(path),
                filename=path.name,
                extension=path.suffix.lower(),
                bytes=path.stat().st_size if path.exists() else None,
                source_posture=source_posture,
                text_chars=len(text),
            ),
            text,
        )
    except Exception as exc:  # pragma: no cover - exercised by live docs.
        return (
            DocumentRecord(
                doc_id=doc_id,
                path=str(path),
                filename=path.name,
                extension=path.suffix.lower(),
                bytes=path.stat().st_size if path.exists() else None,
                source_posture=source_posture,
                load_error=f"{type(exc).__name__}: {exc}",
            ),
            "",
        )


def convert_to_text(path: Path, *, converter: MarkItDown | None = None) -> str:
    if path.suffix.lower() in {".txt", ".md", ".json", ".csv", ".eml"}:
        return path.read_text(encoding="utf-8", errors="replace")
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        generic = (converter or MarkItDown()).convert(path).text_content or ""
        generic = generic.replace("NaN", "")
        structured = workbook_to_text(path)
        if generic.strip():
            return generic + "\n\n# Structured workbook rows\n" + structured
        return structured
    result = (converter or MarkItDown()).convert(path)
    return result.text_content or ""


def normalize_extracted_text(text: str) -> str:
    """Clean common PDF extraction glyph artifacts before retrieval and synthesis."""
    if not text:
        return text
    return text.translate(PDF_EXTRACTION_GLYPH_MAP)


def workbook_to_text(path: Path) -> str:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        parts: list[str] = []
        for sheet in workbook.worksheets:
            parts.append(f"# Workbook: {path.name}")
            parts.append(f"## Sheet: {sheet.title}")
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            for row_index, row in enumerate(rows, 1):
                cells = [format_cell(value) for value in row]
                while cells and cells[-1] == "":
                    cells.pop()
                if not any(cells):
                    continue
                parts.append(f"row {row_index}: " + " | ".join(cells))
        return "\n".join(parts)
    finally:
        workbook.close()


def format_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def chunk_text(
    text: str,
    *,
    doc_id: str,
    chunk_size_chars: int = 4000,
    chunk_overlap_chars: int = 400,
) -> list[ChunkRecord]:
    if chunk_size_chars < 1:
        raise ValueError("chunk_size_chars must be positive")
    if chunk_overlap_chars >= chunk_size_chars:
        raise ValueError("chunk_overlap_chars must be smaller than chunk_size_chars")
    if not text:
        return []

    chunks: list[ChunkRecord] = []
    start = 0
    index = 0
    step = chunk_size_chars - chunk_overlap_chars
    while start < len(text):
        end = min(start + chunk_size_chars, len(text))
        chunk_id = f"{doc_id}_chunk_{index:04d}"
        chunks.append(ChunkRecord(chunk_id=chunk_id, doc_id=doc_id, index=index, text=text[start:end]))
        if end == len(text):
            break
        start += step
        index += 1

    linked: list[ChunkRecord] = []
    for i, chunk in enumerate(chunks):
        linked.append(
            ChunkRecord(
                chunk_id=chunk.chunk_id,
                doc_id=chunk.doc_id,
                index=chunk.index,
                text=chunk.text,
                prev_chunk=chunks[i - 1].chunk_id if i > 0 else None,
                next_chunk=chunks[i + 1].chunk_id if i + 1 < len(chunks) else None,
            )
        )
    return linked
