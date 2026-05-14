from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from irys_harness.benchmarks.harvey_eval import prepare_harvey_eval_package


class HarveyEvalTests(unittest.TestCase):
    def test_prepare_harvey_eval_package_copies_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            artifact = root / "artifact.docx"
            artifact.write_text("content", encoding="utf-8")
            trace = {
                "benchmark": "harvey_lab_sample",
                "run_id": "run123",
                "task_id": "area/task",
                "artifacts": [{"path": str(artifact)}],
            }
            package = prepare_harvey_eval_package(trace, harvey_root=root / "harvey")
            self.assertEqual(package.run_id, "irys/run123")
            self.assertEqual(len(package.copied_files), 1)
            self.assertTrue(Path(package.copied_files[0]).exists())

    def test_prepare_harvey_eval_package_sanitizes_xlsx_cells(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            artifact = root / "matrix.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "A" * 2000
            workbook.save(artifact)
            workbook.close()
            trace = {
                "benchmark": "harvey_lab_sample",
                "run_id": "run123",
                "task_id": "area/task",
                "artifacts": [{"path": str(artifact)}],
            }
            package = prepare_harvey_eval_package(trace, harvey_root=root / "harvey")
            copied = Path(package.copied_files[0])
            copied_workbook = load_workbook(copied, read_only=True)
            try:
                self.assertEqual(len(copied_workbook.active["A1"].value), 500)
            finally:
                copied_workbook.close()
            self.assertEqual(package.validation[0]["mode"], "sanitized_xlsx")


if __name__ == "__main__":
    unittest.main()
