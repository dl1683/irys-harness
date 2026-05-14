from __future__ import annotations

import unittest

from irys_harness.benchmarks.harvey import (
    build_document_comparison_digest,
    build_deterministic_covenant_calculation_digest,
    build_deterministic_checklist_digest,
    build_numeric_fact_digest,
    build_covenant_calculation_worker_prompt,
    build_deliverable_contract,
    build_deliverable_atom_map,
    build_synthesis_prompt,
    build_task_family_digest,
    is_encoded_artifact_answer,
    is_anemic_synthesis_answer,
    build_anemic_synthesis_fallback_answer,
    build_metadata_queries,
    build_numeric_audit_worker_prompt,
    build_provision_comparison_worker_prompt,
    needs_checklist_worker,
    needs_covenant_calculation_worker,
)
from irys_harness.config import load_config
from irys_harness.state import BenchmarkTask, RunState
from openpyxl import Workbook


class HarveyQueryTests(unittest.TestCase):
    def test_queries_include_filenames_not_criteria(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="area/task",
            question="Compare draft petition and checklist.",
            context_files=["/tmp/exhibit-index.docx"],
            answer_schema={"deliverables": ["memo.docx"]},
            metadata={"tags": ["immigration"], "criteria": [{"title": "gold hidden issue"}]},
        )
        queries = build_metadata_queries(task)
        joined = " ".join(queries)
        self.assertIn("exhibit index", joined)
        self.assertNotIn("gold hidden issue", joined)

    def test_provision_worker_prompt_targets_governance_comparison(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="capital/task",
            question="Compare charter against underwriting agreement.",
            context_files=[],
            answer_schema={"deliverables": ["report.docx"]},
            metadata={},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[],
            chunks=[{"doc_id": "doc_1", "chunk_id": "c1", "text": "Section 203 and written consent"}],
        )
        prompt = build_provision_comparison_worker_prompt(state)
        self.assertIn("federal forum", prompt)
        self.assertIn("DGCL Section 203", prompt)
        self.assertIn("stockholder action by written consent", prompt)
        self.assertIn("will provide", prompt)
        self.assertIn("default law", prompt)

    def test_document_comparison_digest_builds_generic_issue_inventory(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="intellectual-property/analyze-counterparty-markup-of-ip-assignment-agreement",
            question="Analyze the counterparty markup of the IP assignment agreement.",
            answer_schema={"deliverables": ["markup-analysis.docx"]},
            metadata={"practice_area": "intellectual-property"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {
                    "doc_id": "doc_1",
                    "filename": "client-form-ip-assignment-agreement.docx",
                    "extension": ".docx",
                },
                {
                    "doc_id": "doc_2",
                    "filename": "counterparty-markup-ip-assignment-agreement.docx",
                    "extension": ".docx",
                },
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Section 4.2 requires assignment of all intellectual property and source code. "
                        "Section 8.1 caps liability at fees paid in the prior 12 months."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c2",
                    "index": 0,
                    "text": (
                        "Counterparty markup deletes the source code assignment language, adds a 30 days cure period, "
                        "and removes the liability cap for indemnification claims."
                    ),
                },
            ],
        )
        contract = build_deliverable_contract(state)
        state.task.answer_schema["deliverable_contract"] = contract
        digest = build_document_comparison_digest(state)
        self.assertEqual(contract["task_family"], "document_comparison")
        self.assertIn("Deterministic document-comparison digest", digest)
        self.assertIn("IP, license, and ownership", digest)
        self.assertIn("Liability, indemnity, and remedies", digest)
        self.assertIn("30 days", digest)
        self.assertIn("counterparty/current markup", digest)
        self.assertIn("Cross-Document Issue Inventory", build_task_family_digest(state))

    def test_covenant_worker_routes_and_requests_formula_workbook(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="banking-finance/task",
            question="Compare borrower covenant compliance certificate.",
            context_files=[],
            answer_schema={"deliverables": ["report.docx"]},
            metadata={"practice_area": "banking-finance"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"filename": "credit-agreement-excerpts.docx"}],
            chunks=[{"doc_id": "doc_1", "chunk_id": "c1", "text": "EBITDA and leverage ratio"}],
        )
        self.assertTrue(needs_covenant_calculation_worker(state))
        prompt = build_covenant_calculation_worker_prompt(state)
        self.assertIn("corrected Total Funded Debt", prompt)
        self.assertIn("conservative EBITDA", prompt)
        self.assertIn("Always show arithmetic", prompt)
        numeric_prompt = build_numeric_audit_worker_prompt(state)
        self.assertIn("Mandatory figures table", numeric_prompt)
        self.assertIn("realized versus projected savings", numeric_prompt)
        self.assertIn("named litigation settlements", numeric_prompt)

    def test_numeric_fact_digest_extracts_relevant_workbook_rows(self) -> None:
        import tempfile
        from pathlib import Path

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "finance.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Pro Forma Cost Savings"
            sheet.append(["Item", "Amount"])
            sheet.append(["Actual savings realized", 1050000])
            sheet.append(["Remaining projected savings", 3150000])
            workbook.save(path)
            workbook.close()
            task = BenchmarkTask(
                benchmark="harvey_lab_sample",
                task_id="banking-finance/task",
                question="Analyze covenant compliance.",
            )
            state = RunState(
                task=task,
                config=load_config(),
                documents=[{"doc_id": "doc_1", "path": str(path)}],
            )
            digest = build_numeric_fact_digest(state)
            self.assertIn("Actual savings realized", digest)
            self.assertIn("Remaining projected savings", digest)

    def test_covenant_digest_distinguishes_unadjusted_and_adjusted_ebitda(self) -> None:
        digest = build_deterministic_covenant_calculation_digest(
            """
            total funded debt; subordinated debt; capital lease obligations; pro forma cost savings; apex settlement;
            capital expenditures.
            Unadjusted EBITDA | $20,790,000.
            Consolidated EBITDA (as adjusted) | $38,750,000.
            Prior compliance certificates added back | $9,800,000.
            current TTM addback of $6,300,000.
            The lifetime restructuring overage is $1,100,000.
            """
        )
        unadjusted_index = digest.find("Borrower's unadjusted EBITDA = $20,790,000")
        adjusted_index = digest.find("Reported Consolidated EBITDA as adjusted before lender corrections = $38,750,000")
        self.assertGreaterEqual(unadjusted_index, 0)
        self.assertGreater(adjusted_index, unadjusted_index)
        self.assertIn("$9,800,000 prior cumulative + $6,300,000 current claimed", digest)
        self.assertIn("$1,100,000 must be excluded", digest)

    def test_checklist_digest_preserves_form_and_exhibit_requirements(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="immigration/task",
            question="Compare filing checklist and draft petition.",
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "text": (
                        "Form I-140, Immigrant Petition for Alien Workers (Ed. 12/23/22). "
                        "Verify current edition on USCIS website before filing; USCIS rejects petitions filed on superseded form editions. "
                        "Exhibit Range (per this checklist) | Exhibit A through Exhibit X (24 total exhibits). "
                        "Checklist Item #87 requires the Field Overview Report / country conditions report as Exhibit X. "
                        "| **Exhibit D** | Expert Letter from Dr. Henrik Johansson, Professor of Neural Engineering | 30-39 | "
                        "| **Exhibit E** | Expert Letter from Dr. Wei-Lin Tsai, Chief Scientific Officer | 40-48 | "
                        "| **Exhibit F** | Expert Letter from Dr. Fatima Al-Rashidi, Director | 49-58 | "
                        "| **Exhibit G** | Expert Letter from Dr. James Moriarty, Department Chair | 59-67 | "
                        "No further exhibits follow Exhibit W. Exhibit H not assigned. "
                        "Dr. Sandra Okonkwo is referenced as Exhibit H but not included."
                    ),
                }
            ],
        )
        digest = build_deterministic_checklist_digest(state)
        self.assertIn("Form I-140 Ed. 12/23/22", digest)
        self.assertIn("USCIS rejects petitions filed on superseded form editions", digest)
        self.assertIn("Exhibits A through X, 24 total exhibits", digest)
        self.assertIn("Exhibit X is the country conditions / field overview report", digest)
        self.assertIn("stops at Exhibit W", digest)
        self.assertIn("Dr. Henrik Johansson (Exhibit D)", digest)
        self.assertIn("Dr. James Moriarty (Exhibit G)", digest)
        self.assertIn("Dr. Sandra Okonkwo is referenced as Exhibit H but is not present", digest)

    def test_deliverable_contract_plans_section_382_workbook_tabs(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="tax/analyze-section-382-analysis",
            question="Prepare section-382-analysis-workbook.xlsx and tax-memorandum.docx.",
            context_files=[],
            answer_schema={"deliverables": ["section-382-analysis-workbook.xlsx", "tax-memorandum.docx"]},
            metadata={"practice_area": "tax"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        workbook = contract["deliverables"][0]
        sheet_names = [sheet["name"] for sheet in workbook["workbook_sheets"]]
        self.assertIn("Shareholder Register", sheet_names)
        self.assertIn("Ownership Shift Calculations", sheet_names)
        self.assertIn("Section 382 Limit Computation", sheet_names)
        self.assertIn("NOL Credit Utilization Impact", sheet_names)
        self.assertEqual(contract["task_family"], "tax_section_382_model")

    def test_deliverable_contract_plans_fund_economics_workbook_tabs(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="funds-asset-management/analyze-fund-economics-comparison",
            question="Prepare fund-economics-comparison.xlsx covering side letters, MFN rights, PPM, and LPA.",
            context_files=[],
            answer_schema={"deliverables": ["fund-economics-comparison.xlsx"]},
            metadata={"practice_area": "funds-asset-management"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        sheet_names = [sheet["name"] for sheet in contract["deliverables"][0]["workbook_sheets"]]
        self.assertIn("Side Letter Economics Matrix", sheet_names)
        self.assertIn("MFN Impact Model", sheet_names)
        self.assertIn("PPM LPA Discrepancy Log", sheet_names)
        self.assertIn("Fund IV to V Comparison", sheet_names)

    def test_package_plan_allocates_fee_letter_and_issues_memo(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="banking-finance/draft-fee-letter",
            question=(
                "Draft a fee letter for a senior secured credit facility and flag cross-document "
                "inconsistencies in a separate issues memo. Output: fee-letter.docx and issues-memorandum.docx."
            ),
            context_files=[],
            answer_schema={"deliverables": ["fee-letter.docx", "issues-memorandum.docx"]},
            metadata={"practice_area": "banking-finance"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        roles = {item["filename"]: item for item in contract["deliverables"]}
        self.assertEqual(contract["package_plan"]["package_kind"], "instrument_plus_issues_package")
        self.assertEqual(roles["fee-letter.docx"]["artifact_role"], "fee_letter")
        self.assertEqual(roles["issues-memorandum.docx"]["artifact_role"], "issues_memo")
        self.assertIn("Ticking fee start date and calculation", roles["fee-letter.docx"]["required_sections"])
        self.assertIn("Cross-document inconsistencies", roles["issues-memorandum.docx"]["required_sections"])

    def test_package_plan_builds_disclosure_schedule_artifacts_without_criteria(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="corporate-ma/draft-disclosure-schedule-preparation",
            question="Prepare the full disclosure schedule package for the UPA.",
            context_files=[],
            answer_schema={
                "deliverables": [
                    "disclosure-schedule-master.docx",
                    "schedule-3-01.docx",
                    "contracts-matrix.xlsx",
                    "employee-census.xlsx",
                ]
            },
            metadata={"practice_area": "corporate-ma", "criteria": [{"title": "hidden scorer atom"}]},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        roles = {item["filename"]: item for item in contract["deliverables"]}
        self.assertEqual(contract["package_plan"]["package_kind"], "disclosure_schedule_package")
        self.assertEqual(roles["disclosure-schedule-master.docx"]["artifact_role"], "disclosure_schedule_master")
        self.assertEqual(roles["schedule-3-01.docx"]["artifact_role"], "disclosure_schedule")
        self.assertIn("Table of contents for schedules 3.1 through 3.26", roles["disclosure-schedule-master.docx"]["required_sections"])
        self.assertIn("Schedule 3.1 heading", roles["schedule-3-01.docx"]["required_sections"])
        contract_sheets = [sheet["name"] for sheet in roles["contracts-matrix.xlsx"]["workbook_sheets"]]
        employee_sheets = [sheet["name"] for sheet in roles["employee-census.xlsx"]["workbook_sheets"]]
        self.assertIn("Material Contract Matrix", contract_sheets)
        self.assertIn("Employee Census", employee_sheets)
        self.assertNotIn("hidden scorer atom", str(contract))

    def test_disclosure_package_accessories_do_not_inherit_tax_memo_sections(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="corporate-ma/draft-disclosure-schedule-preparation",
            question=(
                "Prepare disclosure schedules, seller certificates, consent letters, opinion outline, "
                "data-room mapping, transfer pricing memo, and outstanding items memo. Include tax nexus facts."
            ),
            context_files=[],
            answer_schema={
                "deliverables": [
                    "seller-certificate.docx",
                    "mac-certificate.docx",
                    "kwp-opinion-outline.docx",
                    "data-room-mapping.docx",
                    "transfer-pricing-memo.docx",
                    "landlord-consent-letter.docx",
                    "outstanding-items-memo.docx",
                ]
            },
            metadata={"practice_area": "corporate-ma"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        roles = {item["filename"]: item for item in contract["deliverables"]}

        expected_roles = {
            "seller-certificate.docx": "seller_certificate",
            "mac-certificate.docx": "mac_certificate",
            "kwp-opinion-outline.docx": "opinion_outline",
            "data-room-mapping.docx": "data_room_mapping",
            "transfer-pricing-memo.docx": "transfer_pricing_memo",
            "landlord-consent-letter.docx": "consent_letter",
            "outstanding-items-memo.docx": "outstanding_items_memo",
        }
        for filename, expected_role in expected_roles.items():
            self.assertEqual(roles[filename]["artifact_role"], expected_role)
            self.assertNotIn("Tax issue matrix", roles[filename]["required_sections"])
        self.assertIn("Bringdown representations and warranties", roles["seller-certificate.docx"]["required_sections"])
        self.assertIn("Covered period and no-MAC statement", roles["mac-certificate.docx"]["required_sections"])
        self.assertIn("Requested consent grant", roles["landlord-consent-letter.docx"]["required_sections"])
        self.assertIn("Outstanding item tracker", roles["outstanding-items-memo.docx"]["required_sections"])

    def test_synthesis_prompt_includes_package_plan_and_filename_headings(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="funds-asset-management/draft-lpa-drafting",
            question="Draft a complete LPA, with an issues memo and side letter checklist.",
            context_files=[],
            answer_schema={"deliverables": ["lpa-draft.docx", "issues-memo.docx", "side-letter-checklist.docx"]},
            metadata={"practice_area": "funds-asset-management"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        atom_map = build_deliverable_atom_map(
            contract,
            "\n".join(
                [
                    "Fund IV LPA should use an 8% preferred return and 20% carry.",
                    "Side letter checklist should track MFN and LPAC seat requests.",
                ]
            ),
        )
        state.task.answer_schema["deliverable_contract"] = contract
        state.final_packet = {
            "deliverable_contract": contract,
            "package_plan": contract["package_plan"],
            "deliverable_atom_map": atom_map,
            "cheap_worker_summary": "Fund IV has a 2.0% management fee and side-letter MFN issues.",
            "verified_evidence": [],
        }
        prompt = build_synthesis_prompt(state)
        self.assertIn("Package plan", prompt)
        self.assertIn("Deliverable atom map", prompt)
        self.assertIn("fund_formation_lpa_package", prompt)
        self.assertIn("Side letter checklist should track MFN", prompt)
        self.assertIn("top-level section for every requested filename", prompt)
        self.assertIn("lpa-draft.docx", prompt)
        self.assertIn("side-letter-checklist.docx", prompt)

    def test_deliverable_atom_map_allocates_package_source_lines(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="corporate-ma/draft-disclosure-schedule-preparation",
            question="Prepare disclosure schedules and ancillary closing deliverables.",
            context_files=[],
            answer_schema={
                "deliverables": [
                    "schedule-3-13.docx",
                    "seller-certificate.docx",
                    "transfer-pricing-memo.docx",
                ]
            },
            metadata={"practice_area": "corporate-ma"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        atom_map = build_deliverable_atom_map(
            contract,
            "\n".join(
                [
                    "Schedule 3.13: ISO 9001 certificate expires January 21, 2025 and renewal audit is pending.",
                    "Seller certificate should bring down representations and warranties subject to disclosure schedule exceptions.",
                    "Transfer pricing memo should cover intercompany services, Texas nexus, and tax exposure.",
                ]
            ),
        )
        mapped = atom_map["deliverables"]
        self.assertEqual(mapped["schedule-3-13.docx"]["atom_count"], 1)
        self.assertIn("ISO 9001", mapped["schedule-3-13.docx"]["atoms"][0]["text"])
        self.assertIn("Seller certificate", mapped["seller-certificate.docx"]["atoms"][0]["text"])
        self.assertIn("Transfer pricing", mapped["transfer-pricing-memo.docx"]["atoms"][0]["text"])

    def test_deliverable_atom_map_does_not_match_adjacent_schedule_numbers(self) -> None:
        deliverables = ["schedule-3-02.docx", "schedule-3-20.docx", "seller-certificate.docx"]
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="corporate-ma/draft-disclosure-schedule-preparation",
            question="Prepare the full disclosure schedule package for the UPA.",
            context_files=[],
            answer_schema={"deliverables": deliverables},
            metadata={"practice_area": "corporate-ma"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        atom_map = build_deliverable_atom_map(
            contract,
            "\n".join(
                [
                    "| `doc_0013` | `schedule-3-20.docx` | CGL policy (Reliance National); Products liability sublimit ($1M); D&O policy. |",
                    "Seller certificate should bring down representations and warranties subject to disclosure schedule exceptions.",
                ]
            ),
        )
        mapped = atom_map["deliverables"]
        self.assertEqual(mapped["schedule-3-02.docx"]["atom_count"], 0)
        self.assertEqual(mapped["schedule-3-20.docx"]["atom_count"], 1)
        self.assertIn("Reliance National", mapped["schedule-3-20.docx"]["atoms"][0]["text"])
        self.assertIn("Seller certificate", mapped["seller-certificate.docx"]["atoms"][0]["text"])

    def test_funds_route_does_not_match_unrelated_stipulation_text(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="tax/analyze-counterparty-markup-of-proposed-stipulation-of-facts",
            question="Analyze counterparty markup of proposed stipulation of facts.",
            context_files=[],
            answer_schema={"deliverables": ["tax-memo.docx"]},
            metadata={"practice_area": "tax"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"filename": "fund-iv-lpa-excerpts.docx"}],
        )
        contract = build_deliverable_contract(state)
        self.assertNotEqual(contract["task_family"], "funds_asset_management_review")
        self.assertNotIn("Near-top funds required findings", contract["deliverables"][0]["required_sections"])

    def test_real_estate_digest_routes_and_preserves_lease_issue_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="real-estate/analyze-counterparty-markup-of-commercial-lease-agreement",
            question="Analyze counterparty markup of commercial lease agreement.",
            context_files=[],
            answer_schema={"deliverables": ["redline-analysis-memo.docx"]},
            metadata={"practice_area": "real-estate"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        digest = build_task_family_digest(state)
        self.assertEqual(contract["task_family"], "real_estate_transaction_review")
        self.assertIn("Critical / high / moderate real estate issue tiers", contract["deliverables"][0]["required_sections"])
        self.assertIn("ROFO to ROFR", digest)
        self.assertIn("$686,850", digest)
        self.assertIn("48,200 RSF", digest)
        self.assertIn("Sarah Beckford", digest)
        self.assertIn("litigation-only remedy", digest)

    def test_tax_controversy_digest_preserves_stipulation_issue_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="tax/analyze-counterparty-markup-of-proposed-stipulation-of-facts",
            question="Analyze counterparty markup of proposed stipulation of facts.",
            context_files=[],
            answer_schema={"deliverables": ["stipulation-markup-analysis.docx"]},
            metadata={"practice_area": "tax"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        digest = build_task_family_digest(state)
        self.assertEqual(contract["task_family"], "tax_controversy_review")
        self.assertIn("Tax issue matrix", contract["deliverables"][0]["required_sections"])
        self.assertIn("38.1% of Fund IV equity", digest)
        self.assertIn("$1,221,660", digest)
        self.assertIn("34 IRS-modified paragraphs", digest)
        self.assertIn("12 IRS-added paragraphs", digest)
        self.assertIn("65% of his time", digest)
        self.assertIn("paragraph 157", digest)
        self.assertIn("14.8%", digest)
        self.assertIn("complete qualifier", digest)

    def test_tax_section382_digest_preserves_workbook_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="tax/analyze-section-382-analysis",
            question="Analyze Section 382 ownership changes and produce a workbook.",
            context_files=[],
            answer_schema={"deliverables": ["section-382-analysis-workbook.xlsx", "tax-memorandum.docx"]},
            metadata={"practice_area": "tax"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        digest = build_task_family_digest(state)
        sheet_names = [sheet["name"] for sheet in contract["deliverables"][0]["workbook_sheets"]]
        self.assertIn("Ownership Shift Calculations", sheet_names)
        self.assertIn("Section 382 Limit Computation", sheet_names)
        self.assertIn("NOL Credit Utilization Impact", sheet_names)
        self.assertIn("$14,976,000", digest)
        self.assertIn("$12,500,000", digest)
        self.assertIn("$6,500,000", digest)
        self.assertIn("4,200,000 shares", digest)
        self.assertIn("Section 382 rights plan", digest)

    def test_tax_idr_and_returns_digest_preserves_exact_amounts(self) -> None:
        idr_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="tax/analyze-irs-information-document-request-for-completeness-and-risk-issues",
            question="Analyze an IRS information document request for completeness and risk issues.",
            context_files=[],
            answer_schema={"deliverables": ["tax-issue-memorandum.docx"]},
            metadata={"practice_area": "tax"},
        )
        returns_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="tax/compare-assessed-tax-positions-against-filed-returns",
            question="Compare assessed tax positions against filed returns.",
            context_files=[],
            answer_schema={"deliverables": ["variance-analysis-memo.docx"]},
            metadata={"practice_area": "tax"},
        )
        idr_digest = build_task_family_digest(RunState(task=idr_task, config=load_config(), documents=[]))
        returns_digest = build_task_family_digest(RunState(task=returns_task, config=load_config(), documents=[]))
        self.assertIn("EIN 84-3291057", idr_digest)
        self.assertIn("$365,239", idr_digest)
        self.assertIn("$2,656,025", idr_digest)
        self.assertIn("$2,073,975", idr_digest)
        self.assertIn("$236,667", returns_digest)
        self.assertIn("$2,105,000", returns_digest)
        self.assertIn("$13.32M", returns_digest)
        self.assertIn("$168,000", returns_digest)

    def test_checklist_worker_is_gated_to_relevant_tasks(self) -> None:
        structured_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="structured-finance/task",
            question="Compare offering memorandum against indenture.",
            answer_schema={"deliverables": ["discrepancy-report.docx"]},
            metadata={"practice_area": "structured-finance-securitization"},
        )
        immigration_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="immigration/task",
            question="Compare filing checklist and draft petition exhibits.",
            answer_schema={"deliverables": ["memo.docx"]},
            metadata={"practice_area": "immigration"},
        )
        self.assertFalse(needs_checklist_worker(RunState(task=structured_task, config=load_config())))
        self.assertTrue(needs_checklist_worker(RunState(task=immigration_task, config=load_config())))

    def test_structured_finance_digest_preserves_closing_checklist_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="structured-finance-securitization/compare-closing-checklist-against-transaction-documents",
            question="Compare closing checklist against transaction documents.",
            answer_schema={"deliverables": ["closing-checklist-report.docx"]},
            metadata={"practice_area": "structured-finance-securitization"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "closing-checklist.docx"}],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Lakeshore Auto Receivables Trust 2024-2 Master Closing Checklist "
                        "item 4.1 says Transfer and Servicing Agreement. SSA Section 6.04 "
                        "requires a Back-Up Servicing Agreement. Item 10.1 rating agency "
                        "letter date, item 5.1 Class C Notes, item 11.2 UCC filing, and "
                        "item 10.3 10b-5 negative assurance letter are under review."
                    ),
                }
            ],
        )
        contract = build_deliverable_contract(state)
        digest = build_task_family_digest(state)
        self.assertEqual(contract["task_family"], "structured_finance_securitization_review")
        self.assertIn("Transfer and Servicing Agreement", digest)
        self.assertIn("Sale and Servicing Agreement", digest)
        self.assertIn("Back-Up Servicing Agreement", digest)
        self.assertIn("SSA Section 6.04", digest)
        self.assertIn("Class C Notes", digest)
        self.assertIn("10b-5 negative assurance", digest)
        self.assertIn("Delaware", digest)

    def test_structured_finance_digest_preserves_om_and_clo_calculations(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="structured-finance-securitization/compare-offering-memorandum-against-indenture",
            question="Compare Offering Memorandum against Indenture and analyze Trident indenture markup.",
            answer_schema={"deliverables": ["comparison-report.docx"]},
            metadata={"practice_area": "structured-finance-securitization"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "indenture.docx"}],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Crescent Auto Receivables Trust Rule 144A Minimum OC Amount "
                        "Reserve Account Required Balance. Whitmore CLO 2025-3 Trident "
                        "Institutional Partners Ashford Hale markup includes Discount Obligation, "
                        "Interest Diversion Test, clean-up call, refinancing cost cap, and "
                        "Weighted Average Life provisions."
                    ),
                }
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("6.00% vs 5.75%", digest)
        self.assertIn("$3,437,500", digest)
        self.assertIn("31,412 vs ~31,200", digest)
        self.assertIn("2.0% annual cap", digest)
        self.assertIn("$8.5M", digest)
        self.assertIn("$425M", digest)
        self.assertIn("30-50 bps", digest)

    def test_structured_finance_digest_preserves_psa_and_collateral_tape_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="structured-finance-securitization/compare-collateral-tape-against-eligibility-criteria",
            question="Review pooling and servicing agreement markup and collateral tape.",
            answer_schema={"deliverables": ["eligibility-report.docx"]},
            metadata={"practice_area": "structured-finance-securitization"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "collateral-tape-2025-06-27.xlsx"}],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Whitmore Auto Receivables Trust 2025-1 Ridgefield National Bank "
                        "backup servicer markup. Thornfield CLO collateral tape includes Loan #14, "
                        "Loan #41, Loan #58, Loan #63, Industry #18 and Caa1 concentration review."
                    ),
                }
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("three-month rolling average", digest)
        self.assertIn("5.00%", digest)
        self.assertIn("Loan #41", digest)
        self.assertIn("$10M", digest)
        self.assertIn("Loan #63", digest)
        self.assertIn("Industry #18", digest)
        self.assertIn("$31.875M", digest)

    def test_white_collar_digest_preserves_dpa_payment_and_monitor_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="white-collar-defense-investigations/analyze-counterparty-markup-of-deferred-prosecution-agreement",
            question="Analyze the counterparty markup of the Deferred Prosecution Agreement.",
            answer_schema={"deliverables": ["dpa-markup-report.docx"]},
            metadata={"practice_area": "white-collar-defense-investigations"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "govt-counter-markup.docx"}],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Vantage Meridian Holdings VMH Deferred Prosecution Agreement "
                        "transmittal references Attachments A, B, and C. Section 7(c) "
                        "monitor access, Institutional Knowledge, Compliance Certification, "
                        "70% and 30% payment schedule, and 18 U.S.C. Section 1349 are at issue."
                    ),
                }
            ],
        )
        contract = build_deliverable_contract(state)
        digest = build_task_family_digest(state)
        self.assertEqual(contract["task_family"], "white_collar_investigations_review")
        self.assertIn("Missing Attachment C", digest)
        self.assertIn("$92.5M credit", digest)
        self.assertIn("$274.75M and $117.75M", digest)
        self.assertIn("$436.8M original government penalty", digest)
        self.assertIn("reasonably related business-unit", digest)
        self.assertIn("Debarment waiver and federal-contracting risk", digest)
        self.assertIn("Institutional Knowledge", digest)
        self.assertIn("Section 1349", digest)

    def test_white_collar_digest_preserves_subpoena_production_and_retention_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="white-collar-defense-investigations/compare-federal-grand-jury-subpoena-scope-against-corporate-document-retention-policy",
            question="Compare federal grand jury subpoena scope against corporate document retention policy.",
            answer_schema={"deliverables": ["subpoena-retention-report.docx"]},
            metadata={"practice_area": "white-collar-defense-investigations"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "document-retention-policy.docx"}],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Grand Jury Subpoena Hargrove Specialty Clearpath retention policy "
                        "Calverley audit Lisa Egan Ironvault Baton Rouge consent decree "
                        "DMR NPDES Category A Category B Category N March 17 2025."
                    ),
                }
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("Subpoena Category Letter Checklist", digest)
        self.assertIn("Clearpath 2019 internal lab data", digest)
        self.assertIn("GJ-2025-04418", digest)
        self.assertIn("Microsoft Teams pre-February 2022 loss", digest)
        self.assertIn("Post-hold destruction of two boxes", digest)
        self.assertIn("March 17, 2025", digest)
        self.assertIn("Lisa Egan", digest)
        self.assertIn("Baton Rouge consent decree", digest)
        self.assertIn("regardless-of-date", digest)

    def test_white_collar_digest_preserves_production_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="white-collar-defense-investigations/compare-document-production-set-against-subpoena-request-categories",
            question="Compare document production set against subpoena request categories.",
            answer_schema={"deliverables": ["production-gap-report.docx"]},
            metadata={"practice_area": "white-collar-defense-investigations"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "production-tracker.xlsx"}],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Ridgeline Capital MedCore RC-PROD production tracker "
                        "Category 1 Category 3 Category 5 Category 9 Category 11 Category 15 Category 16 "
                        "Krell interview Adler Connolly Webb email multi-category tagging."
                    ),
                }
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("52,318 total documents", digest)
        self.assertIn("Category 1 has about 2,134 documents", digest)
        self.assertIn("Category 5 has about 2,847 documents", digest)
        self.assertIn("Category 9 zero-production problem", digest)
        self.assertIn("Category 16 shows zero documents", digest)
        self.assertIn("Krell interview and Adler Connolly communications", digest)

    def test_white_collar_digest_preserves_statutory_and_sec_communication_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="white-collar-defense-investigations/compare-investigation-memorandum-against-applicable-statutes",
            question="Compare investigation memorandum against applicable statutes and SEC referral notice communications.",
            answer_schema={"deliverables": ["statutory-gap-report.docx"]},
            metadata={"practice_area": "white-collar-defense-investigations"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "investigation-memorandum.docx"}],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Bellweather Holt Whitmore Capital Management Section 206(2) "
                        "Rule 10b-5 206(4)-5 1512 1513 2B1.1 SEC referral notice "
                        "RidgeChat Delacroix dark pools catalyst per your instruction."
                    ),
                }
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("Section 206(2) requires only negligence", digest)
        self.assertIn("$350/$150", digest)
        self.assertIn("18 U.S.C. Sections 1512/1513", digest)
        self.assertIn("USSG Section 2B1.1", digest)
        self.assertIn("Tyrell did file with the SEC", digest)
        self.assertIn("ERISA/state pension fiduciary obligations", digest)
        self.assertIn("lookback period and the forward ban", digest)

    def test_white_collar_digest_preserves_sec_communication_rows_without_generic_contamination(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="white-collar-defense-investigations/compare-employee-communications-against-sec-referral-notice",
            question="Compare employee communications against SEC referral notice.",
            answer_schema={"deliverables": ["sec-communications-gap-report.docx"]},
            metadata={"practice_area": "white-collar-defense-investigations"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "sec-referral-notice.docx"}],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "SEC referral notice RidgeChat Delacroix Hale Yoon Breslin catalyst "
                        "dark pools per your instruction Ferrante 14 calls 22 texts sure thing."
                    ),
                }
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("April 5 Hale instruction", digest)
        self.assertIn("May 20 NSBM wall-crossed short instruction", digest)
        self.assertIn("March 17 Breslin per your instruction message", digest)
        self.assertIn("14 calls and 22 text messages", digest)
        self.assertIn("Required categorization framework", digest)
        self.assertNotIn("Category-by-Category Production Coverage", digest)

    def test_fund_economics_digest_builds_lp_level_rows(self) -> None:
        import tempfile
        from pathlib import Path

        with tempfile.TemporaryDirectory() as tmp:
            commitment_path = Path(tmp) / "investor-commitment-summary.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Investor Commitment Summary"
            sheet.append(["Investor Name", "Short Name", "Commitment Amount ($)"])
            sheet.append(["Nordhaven Sovereign Wealth Fund", "Nordhaven SWF", 250000000])
            workbook.save(commitment_path)
            workbook.close()
            task = BenchmarkTask(
                benchmark="harvey_lab_sample",
                task_id="funds-asset-management/analyze-fund-economics-comparison",
                question="Analyze fund economics, side letters, MFN rights, PPM and LPA.",
                answer_schema={"deliverables": ["side-letter-economics-matrix.xlsx"]},
                metadata={"practice_area": "funds-asset-management"},
            )
            state = RunState(
                task=task,
                config=load_config(),
                documents=[
                    {
                        "doc_id": "doc_1",
                        "filename": "side-letter-nordhaven-swf.docx",
                        "path": str(Path(tmp) / "side-letter-nordhaven-swf.docx"),
                    },
                    {
                        "doc_id": "doc_2",
                        "filename": "investor-commitment-summary.xlsx",
                        "path": str(commitment_path),
                    },
                ],
                chunks=[
                    {
                        "doc_id": "doc_1",
                        "chunk_id": "c1",
                        "index": 0,
                        "text": (
                            "During the Investment Period, the Management Fee shall be 1.75% per annum. "
                            "Following the Investment Period, the Post-Investment Period Fee shall be 1.25%. "
                            "Carried Interest Reduction: reduced carried interest rate of fifteen percent (15%). "
                            "Preferred Return at the rate of eight percent (8%) per annum, compounded annually."
                        ),
                    }
                ],
            )
            digest = build_task_family_digest(state)
            self.assertIn("Side Letter Economics Matrix", digest)
            self.assertIn("Nordhaven SWF", digest)
            self.assertIn("$250,000,000", digest)
            self.assertIn("1.75%", digest)
            self.assertIn("1.25%", digest)
            self.assertIn("15%", digest)
            self.assertIn("Fund Economics Residual Specificity Packet", digest)
            self.assertIn("Great Lakes has a 9% hurdle", digest)
            self.assertIn("annual reduction is at least about $325K", digest)

    def test_funds_digest_preserves_lpa_markup_specificity_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="funds-asset-management/analyze-counterparty-markup-of-limited-partnership-agreement",
            question="Review the CRPS markup of the Fund III LPA against the GP form and playbook.",
            answer_schema={"deliverables": ["redline-review-memo.docx"]},
            metadata={"practice_area": "funds-asset-management"},
        )
        state = RunState(task=task, config=load_config(), documents=[], chunks=[])
        digest = build_task_family_digest(state)
        self.assertIn("CRPS proposed commitment is $100M", digest)
        self.assertIn("1.50%", digest)
        self.assertIn("17.5%", digest)
        self.assertIn("$14.12M", digest)
        self.assertIn("50% and shortens escrow period from 3 years to 18 months", digest)

    def test_funds_digest_preserves_investment_advisory_specificity_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="funds-asset-management/analyze-counterparty-markup-of-investment-advisory-agreement",
            question="Review the redlined investment advisory agreement against the standard form and playbook.",
            answer_schema={"deliverables": ["redline-review-memorandum.docx"]},
            metadata={"practice_area": "funds-asset-management"},
        )
        state = RunState(task=task, config=load_config(), documents=[], chunks=[])
        digest = build_task_family_digest(state)
        self.assertIn("Fiduciary duty contradiction", digest)
        self.assertIn("$150K annually", digest)
        self.assertIn("$513,750", digest)
        self.assertIn("$1,475,000", digest)

    def test_funds_digest_preserves_side_letter_and_transfer_specificity_rows(self) -> None:
        side_letter = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="funds-asset-management/analyze-counterparty-markup-of-side-letter",
            question="Review the CalSEPS redline side letter.",
            answer_schema={"deliverables": ["calseps-side-letter-analysis-memo.docx"]},
            metadata={"practice_area": "funds-asset-management"},
        )
        transfer = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="funds-asset-management/analyze-counterparty-markup-of-limited-partnership-interest-transfer-agreement",
            question="Analyze the limited partnership interest transfer agreement.",
            answer_schema={"deliverables": ["transfer-review-memo.docx"]},
            metadata={"practice_area": "funds-asset-management"},
        )
        side_digest = build_task_family_digest(RunState(task=side_letter, config=load_config()))
        transfer_digest = build_task_family_digest(RunState(task=transfer, config=load_config()))
        self.assertIn("$875,000", side_digest)
        self.assertIn("9.46%", side_digest)
        self.assertIn("$2,187,500", side_digest)
        self.assertIn("quarterly ESG reporting", side_digest)
        self.assertIn("fee advancement", side_digest)
        self.assertIn("$35,512,000", transfer_digest)
        self.assertIn("$1,325,000", transfer_digest)
        self.assertIn("Northbridge Valuation Services", transfer_digest)

    def test_ipo_charter_digest_preserves_governance_issue_families(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="capital-markets/compare-charter-against-offering",
            question="Compare current charter against underwriting agreement and prospectus for IPO.",
            answer_schema={"deliverables": ["charter-offering-deviation-report.docx"]},
            metadata={"practice_area": "capital-markets"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "underwriting-agreement.docx"}],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Target IPO charter must include a federal forum provision for Securities Act claims, "
                        "prohibit stockholder action by written consent, prohibit cumulative voting, and eliminate "
                        "Series A Preferred, Series B Preferred, Series C Preferred and all anti-dilution preferential rights."
                    ),
                }
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("Deterministic IPO charter comparison digest", digest)
        self.assertIn("federal forum", digest)
        self.assertIn("written-consent permission", digest)
        self.assertIn("Series A/B/C preferred", digest)

    def test_bankruptcy_distribution_digest_preserves_class_and_calculation_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="bankruptcy-restructuring/compare-distribution-amounts-against-plan-requirements",
            question="Review the Q1 distribution report against the confirmed plan and related case documents.",
            answer_schema={"deliverables": ["distribution-compliance-memo.docx"]},
            metadata={"practice_area": "bankruptcy-restructuring"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "confirmed-plan-of-reorganization.docx"},
                {"doc_id": "doc_2", "filename": "q1-distribution-report.xlsx"},
                {"doc_id": "doc_3", "filename": "westlake-stipulation-order.docx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "The confirmed plan requires a class-by-class distribution analysis. "
                        "Classes 1 through 6 receive specified cash distributions, while Classes 7 and 8 "
                        "receive no distribution. Class 1 priority tax claims accrue interest under 11 USC "
                        "1129(a)(9)(C) at 5.25%. Class 5 includes PBGC and non-PBGC employee wage claims."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c2",
                    "index": 0,
                    "text": (
                        "Q1 Distribution Report row 14: Class 6 GUC cash pool $8,500,000; Plan Agent fee "
                        "$37,500; disputed claims reserve $965,473; Tranche 1 reported $4,862,527; "
                        "calculated net Tranche 1 $4,483,216; Tranche 2 reported $3,400,000; total allowed "
                        "Class 6 distributions $49,460,000. The disputed claims reserve must be held in a "
                        "segregated account."
                    ),
                },
                {
                    "doc_id": "doc_3",
                    "chunk_id": "c3",
                    "index": 0,
                    "text": (
                        "Westlake undeliverable funds must be redistributed within 14 days after the "
                        "mailing deadline; the status report shows actual distribution on May 20."
                    ),
                },
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("Deterministic bankruptcy distribution compliance digest", digest)
        self.assertIn("Class-By-Class Distribution Inventory", digest)
        self.assertIn("Class 1", digest)
        self.assertIn("Class 6", digest)
        self.assertIn("Class 8", digest)
        self.assertIn("$8,500,000", digest)
        self.assertIn("$37,500", digest)
        self.assertIn("$965,473", digest)
        self.assertIn("High-Priority Bankruptcy Calculation Checklist", digest)
        self.assertIn("$7,497,027", digest)
        self.assertIn("$2,998,811", digest)
        self.assertIn("$4,483,216", digest)
        self.assertIn("$49,460,000", digest)
        self.assertIn("segregated account", digest)
        self.assertIn("Westlake", digest)
        self.assertIn("within 14 days", digest)

    def test_synthesis_prompt_forbids_encoded_artifacts(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="area/task",
            question="Prepare memo. Output: `memo.docx`.",
            answer_schema={"deliverables": ["memo.docx"]},
            metadata={},
        )
        state = RunState(task=task, config=load_config(), final_packet={"verified_evidence": []})
        prompt = build_synthesis_prompt(state)
        self.assertIn("Do not output base64", prompt)
        self.assertIn("The harness will render your plain text", prompt)
        self.assertIn("bankruptcy distribution compliance deliverables", prompt)
        self.assertNotIn("High-Priority Prenuptial Asset-Rights Matrix", prompt)
        self.assertNotIn("High-Priority Environmental Indemnity Matrix", prompt)
        self.assertNotIn("High-Priority Technology Contract Amendment Deviation Matrix", prompt)
        self.assertNotIn("High-Priority Technology/Data Agreement Clause-Delta Matrix", prompt)
        self.assertTrue(is_encoded_artifact_answer("```xml\n<base64_file><content>" + ("A" * 5000)))
        self.assertFalse(is_encoded_artifact_answer("Executive summary\n\nClass 1 is underpaid."))

    def test_flsa_gap_digest_preserves_threshold_counts_and_position_risks(self) -> None:
        import tempfile
        from pathlib import Path

        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "employee-classification-data.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Employee Data"
            sheet.append(
                [
                    "Employee ID",
                    "Last Name",
                    "First Name",
                    "Property Name",
                    "State",
                    "Department",
                    "Job Title",
                    "Exemption Type",
                    "Annual Salary",
                    "Total Annual Compensation",
                    "Weekly Salary Equivalent",
                    "Date of Last Classification Review",
                    "Notes",
                ]
            )
            rows = [
                ("Guest Services Manager", "Admin", 41600, "WA", 9),
                ("Marketing Specialist", "Administrative", 39500, "WA", 5),
                ("Revenue Analyst", "Professional", 47500, "OR", 4),
                ("Asst Food & Beverage Manager", "Executive", 42900, "ID", 8),
                ("Events Coordinator", "Admin", 44800, "WA", 6),
                ("IT Support Lead", "Prof", 52000, "WA", 4),
                ("Director of Operations", "HCE", 115000, "WA", 5),
            ]
            counter = 1
            for title, exemption, salary, state_code, count in rows:
                for _ in range(count):
                    sheet.append(
                        [
                            f"CMHG-{counter:05d}",
                            "Last",
                            "First",
                            "Property",
                            state_code,
                            "Dept",
                            title,
                            exemption,
                            salary,
                            salary,
                            round(salary / 52, 2),
                            "09/15/2021",
                            "",
                        ]
                    )
                    counter += 1
            summary = workbook.create_sheet("Summary Pivot")
            summary.append(["Salary Band", "Executive", "Administrative", "Professional", "HCE", "Total"])
            summary.append(["STANDARD EAP SALARY THRESHOLD ANALYSIS", None, None, None, None, None])
            summary.append([None, None, None, None, None, None])
            summary.append(["Salary Band", "Executive", "Administrative", "Professional", "HCE", "Total"])
            summary.append(["Below $43,888", 6, 15, 2, 0, 23])
            summary.append(["$43,888 - $58,655", 19, 48, 7, 0, 74])
            summary.append(["$58,656 and above", 87, 76, 13, 14, 190])
            summary.append(["TOTAL", 112, 139, 22, 14, 287])
            summary.append([None, None, None, None, None, None])
            summary.append(["HCE COMPENSATION THRESHOLD ANALYSIS", None, None, None, None, None])
            summary.append([None, None, None, None, None, None])
            summary.append(["Compensation Band", "Count", None, None, None, None])
            summary.append(["Below $132,964 (Fail Phase 1 HCE)", 8, None, None, None, None])
            summary.append(["$132,964 - $151,163 (Pass Phase 1, Fail Phase 2 HCE)", 3, None, None, None, None])
            summary.append(["$151,164 and above (Pass Both HCE Thresholds)", 3, None, None, None, None])
            summary.append(["Total HCE Employees", 14, None, None, None, None])
            summary.append([None, None, None, None, None, None])
            summary.append(["EMPLOYEE COUNT BY STATE", None, None, None, None, None])
            summary.append([None, None, None, None, None, None])
            summary.append(["State", "Count", None, None, None, None])
            summary.append(["WA", 168, None, None, None, None])
            summary.append(["OR", 82, None, None, None, None])
            summary.append(["ID", 37, None, None, None, None])
            summary.append(["Total", 287, None, None, None, None])
            workbook.save(workbook_path)
            workbook.close()
            task = BenchmarkTask(
                benchmark="harvey_lab_sample",
                task_id="corporate-governance/analyze-flsa-overtime-rule-gap-against-current-employee-classifications",
                question="Produce an FLSA overtime gap analysis memo.",
                answer_schema={"deliverables": ["flsa-gap-analysis-memo.docx"]},
                metadata={"practice_area": "corporate-governance"},
            )
            state = RunState(
                task=task,
                config=load_config(),
                documents=[{"doc_id": "doc_1", "filename": workbook_path.name, "path": str(workbook_path)}],
                chunks=[
                    {
                        "doc_id": "doc_2",
                        "chunk_id": "c1",
                        "index": 0,
                        "text": (
                            "Job Title: Guest Services Manager FLSA Status: Exempt Exemption Type: Administrative "
                            "Number of Positions Company-wide: 9 Position Summary: routine operational front-desk work. "
                            "Job Title: Events Coordinator FLSA Status: Exempt Exemption Type: Administrative "
                            "Number of Positions Company-wide: 6 Position Summary: standard event packages and pricing. "
                            "At-risk exempt employees work an average of 46.5 hours per week, generating 6.5 overtime hours weekly. "
                            "Group 1 exposure is $224,824 and Group 2 exposure is $923,712. "
                            "Oakvale Barker identified 11 positions as potentially misclassified; 7 were reclassified and 4 retained. "
                            "Authorities include 29 U.S.C. § 216(b), 29 U.S.C. § 255(a), and 29 C.F.R. § 541.400."
                        ),
                    }
                ],
            )
            digest = build_task_family_digest(state)
            self.assertIn("Deterministic FLSA overtime gap digest", digest)
            self.assertIn("Executive 112", digest)
            self.assertIn("$684/week", digest)
            self.assertIn("$35,568/year", digest)
            self.assertIn("cost-effective alternative", digest)
            self.assertIn("Administrative 139", digest)
            self.assertIn("Professional/Learned 22", digest)
            self.assertIn("Washington 168", digest)
            self.assertIn("Guest Services Manager", digest)
            self.assertIn("| Guest Services Manager | 9 |", digest)
            self.assertIn("Marketing Specialist", digest)
            self.assertIn("Revenue Analyst", digest)
            self.assertIn("| Events Coordinator | 6 |", digest)
            self.assertIn("6.5 overtime hours/week", digest)
            self.assertIn("$1,148,536", digest)
            prompt = build_synthesis_prompt(state)
            self.assertIn("For FLSA overtime gap-analysis deliverables", prompt)
            self.assertIn("Position-Level Classification Risk Matrix", prompt)
            self.assertIn("equivalent experience", digest)
            self.assertIn("29 U.S.C. § 216(b)", digest)
            self.assertIn("29 C.F.R. § 541.400", digest)

    def test_employment_labor_digest_preserves_complaint_claim_defenses(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="employment-labor/analyze-iss-employment-complaint",
            question="Analyze an employment complaint asserting SOX, FLSA, Title VII, FMLA, and Texas IIED claims.",
            answer_schema={"deliverables": ["employment-complaint-analysis.docx"]},
            metadata={"practice_area": "employment-labor"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "complaint.docx"}],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": "SOX OSHA FLSA Title VII front pay outside sales Texas IIED FMLA Greenleaf SEC tip line.",
                }
            ],
        )
        contract = build_deliverable_contract(state)
        digest = build_task_family_digest(state)
        prompt = build_synthesis_prompt(state)
        self.assertEqual(contract["task_family"], "employment_labor_review")
        self.assertIn("SOX administrative exhaustion", digest)
        self.assertIn("SEC tip-line reporting does not substitute for OSHA filing", digest)
        self.assertIn("Title VII statutory cap", digest)
        self.assertIn("$200,000 cap", digest)
        self.assertIn("outside sales has no minimum salary requirement", digest)
        self.assertIn("145 weeks", digest)
        self.assertIn("$347,648", digest)
        self.assertIn("Kevin Stanhope", digest)
        self.assertIn("double hearsay", digest)
        self.assertIn("recommend moving to dismiss the SOX claim", digest)
        self.assertIn("For employment, labor, ADA/FMLA", prompt)

    def test_employment_labor_digest_preserves_executive_and_termination_rows(self) -> None:
        executive_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="employment-labor/analyze-counterparty-markup-of-executive-employment-agreement",
            question="Analyze counterparty markup of executive employment agreement.",
            answer_schema={"deliverables": ["executive-employment-markup-report.docx"]},
            metadata={"practice_area": "employment-labor"},
        )
        executive_state = RunState(
            task=executive_task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "executive-employment-agreement.docx"}],
            chunks=[{"doc_id": "doc_1", "chunk_id": "c1", "index": 0, "text": "Good Reason 280G AB 1076 SB 699 California bonus Board arbitration signing bonus."}],
        )
        executive_digest = build_task_family_digest(executive_state)
        self.assertIn("Bonus target increase", executive_digest)
        self.assertIn("30 months + 2.5x target bonus", executive_digest)
        self.assertIn("California AB 1076 / SB 699", executive_digest)
        self.assertIn("280G gross-up", executive_digest)
        self.assertIn("$468,750 to $700,000", executive_digest)
        self.assertIn("prior-year metrics if no agreement by March 31", executive_digest)
        self.assertIn("Fund III and Fund IV companies use best-net cutbacks", executive_digest)
        self.assertIn("December 31, 2029", executive_digest)

        termination_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="employment-labor/assess-legal-risk-of-proposed-employee-termination",
            question="Assess legal risk of proposed employee termination.",
            answer_schema={"deliverables": ["termination-risk-memo.docx"]},
            metadata={"practice_area": "employment-labor"},
        )
        termination_state = RunState(
            task=termination_task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "termination-record.docx"}],
            chunks=[{"doc_id": "doc_1", "chunk_id": "c1", "index": 0, "text": "Kessler fresh energy different era PIP Kolb Okafor OWBPA OSHA Illinois non-compete."}],
        )
        termination_digest = build_task_family_digest(termination_state)
        self.assertIn("Age-coded comments", termination_digest)
        self.assertIn("Kolb and Okafor", termination_digest)
        self.assertIn("OWBPA requirements", termination_digest)
        self.assertIn("Illinois Freedom to Work Act", termination_digest)
        self.assertIn("Jennifer Kolb, age 34, White female", termination_digest)
        self.assertIn("David Okafor, age 41, Black male", termination_digest)
        self.assertIn("felony plea/conviction", termination_digest)
        self.assertIn("Operations Leader of the Year", termination_digest)

    def test_employment_labor_digest_preserves_accommodation_and_classification_rows(self) -> None:
        accommodation_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="employment-labor/analyze-reasonable-accommodation-request-under-ada-requirements",
            question="Analyze reasonable accommodation request under ADA requirements.",
            answer_schema={"deliverables": ["ada-accommodation-memo.docx"]},
            metadata={"practice_area": "employment-labor"},
        )
        accommodation_state = RunState(
            task=accommodation_task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "accommodation-request.docx"}],
            chunks=[{"doc_id": "doc_1", "chunk_id": "c1", "index": 0, "text": "RRMS Robles March 1 FMLA Ohio Okonkwo performance medical confidentiality."}],
        )
        accommodation_digest = build_task_family_digest(accommodation_state)
        self.assertIn("RRMS disability qualification", accommodation_digest)
        self.assertIn("Robles March 1 stereotyping email", accommodation_digest)
        self.assertIn("Ohio Revised Code Chapter 4112", accommodation_digest)
        self.assertIn("FMLA intersection", accommodation_digest)
        self.assertIn("internal safety practice, not an OSHA mandate", accommodation_digest)
        self.assertIn("Dallas DC forklift exemption", accommodation_digest)
        self.assertIn("FY2024 revenue $1.24B", accommodation_digest)
        self.assertIn("Nashville sit-stand desk", accommodation_digest)

        classification_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="employment-labor/assess-worker-classification-for-proposed-engineering-engagement",
            question="Assess worker classification for proposed engineering engagement.",
            answer_schema={"deliverables": ["classification-memo.docx"]},
            metadata={"practice_area": "employment-labor"},
        )
        classification_state = RunState(
            task=classification_task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "engineering-engagement.docx"}],
            chunks=[{"doc_id": "doc_1", "chunk_id": "c1", "index": 0, "text": "independent contractor laptop VPN work-for-hire Section 16600 economic reality VP of Engineering FICA SDI SUI."}],
        )
        classification_digest = build_task_family_digest(classification_state)
        self.assertIn("Company laptop and VPN", classification_digest)
        self.assertIn("California Business and Professions Code Section 16600", classification_digest)
        self.assertIn("DOL economic-reality factors", classification_digest)
        self.assertIn("FICA approximation", classification_digest)
        self.assertIn("Principal Engineer - Power Electronics", classification_digest)
        self.assertIn("17 U.S.C. 101", classification_digest)
        self.assertIn("$14,073 per year", classification_digest)
        self.assertIn("IT support and marketing vendors", classification_digest)

    def test_eu_distribution_risk_digest_preserves_cross_border_issue_matrix(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="international-trade-sanctions/analyze-cross",
            question="Draft a risk memorandum for the proposed exclusive EU distribution agreement.",
            answer_schema={"deliverables": ["risk-memorandum.docx"]},
            metadata={"practice_area": "international-trade-sanctions"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "draft-distribution-agreement-v4.2.docx", "path": "draft-distribution-agreement-v4.2.docx"},
                {"doc_id": "doc_2", "filename": "falkenberg-regulatory-report.docx", "path": "falkenberg-regulatory-report.docx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Section 7.3 Resale Pricing requires retail pricing within ten percent of MSRP. "
                        "Article 101 TFEU and Regulation (EU) 2022/720 appear elsewhere in the agreement. "
                        "Section 6.2 permits Sub-Distributors without prior written consent. "
                        "Section 11.2 makes Nordlicht responsible for regulatory approvals; Section 11.4 gives Brightwell a compliance warranty. "
                        "Section 14 says each party shall comply with GDPR. "
                        "Section 4.1 gives Brightwell sole discretion over labeling. "
                        "Sections 17.1 and 17.2 choose Texas law and Austin arbitration."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c2",
                    "index": 1,
                    "text": (
                        "Friedrich Wendt sent a March 3, 2025 cost-sharing email for EUR 340,000. "
                        "On May 20, 2025 he disclosed a French competition authorities investigation. "
                        "Brightwell has no EUIPO or Madrid Protocol filings. "
                        "The three BioEnhance patents have no European Patent Office filings, no PCT applications, and the Paris Convention priority period expired. "
                        "Health claims non-compliance affects 9 of 12 SKUs. "
                        "VitaEdge Neuro Focus contains Ashwagandha Root Extract (600 mg). "
                        "VitaEdge Iron Boost contains 65 mg elemental iron per serving, above EFSA 25 mg/day guidance. "
                        "VitaEdge Immune+ contains elderberry extract."
                    ),
                },
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("High-Priority EU Distribution Risk Matrix", digest)
        self.assertIn("RPM / resale pricing", digest)
        self.assertIn("Article 101 TFEU", digest)
        self.assertIn("VBER", digest)
        self.assertIn("Section 11.2", digest)
        self.assertIn("Section 11.4", digest)
        self.assertIn("GDPR", digest)
        self.assertIn("March 3, 2025", digest)
        self.assertIn("May 20, 2025", digest)
        self.assertIn("elderberry", digest)
        self.assertIn("25 mg/day", digest)
        self.assertIn("9 of 12 SKUs", digest)
        self.assertIn("no corresponding European Patent Office filings", digest)
        self.assertIn("only contractual", digest)
        self.assertIn("commercial agent reclassification", digest)

    def test_insurance_coverage_digest_preserves_claim_line_arithmetic(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="insurance/analyze-property-damage-claim-against-commercial-policy-exclusions",
            question="Prepare a coverage determination memorandum for the property damage claim under the commercial policy.",
            answer_schema={"deliverables": ["coverage-determination-memo.docx"]},
            metadata={"practice_area": "insurance"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "policy-sai-cpp-2024-07831.docx", "path": "policy-sai-cpp-2024-07831.docx"},
                {"doc_id": "doc_2", "filename": "calverley-proof-of-loss.docx", "path": "calverley-proof-of-loss.docx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Policy Number: SAI-CPP-2024-07831. Covered Cause of Loss means all risks of direct physical loss "
                        "unless excluded; this is Special Form all-risk coverage. EB-100 excludes equipment not maintained "
                        "according to manufacturer recommendations. FL-200 excludes process fluids. MF-300 has a $150,000 "
                        "mold sublimit and requires reporting within 30 days. Ordinance or Law Coverage has a $500,000 sublimit. "
                        "Business Income has a 72-hour waiting period. Exclusion E has an ensuing loss clause."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c2",
                    "index": 1,
                    "text": (
                        "The January 14, 2025 proof of loss totals $4,730,000. Building Damage is $1,245,000, including "
                        "roof deck and structural beam repair $410,000, electrical system replacement $385,000, office wing "
                        "water/mold damage $215,000, breezeway repair $85,000, and fabrication hall floor coating $150,000. "
                        "Business Personal Property is $1,612,000, including CNC plasma cutting table #3 $485,000 and raw steel "
                        "inventory $340,000. Environmental remediation is $287,500 for Hexacoat 7200 RCRA-regulated material. "
                        "Business Income Loss is $1,178,000, including $890,000 lost revenue and $288,000 extra expense over 47 days. "
                        "Mold remediation is $262,500. Code upgrades are $145,000."
                    ),
                },
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("High-Priority Insurance Coverage Matrix", digest)
        self.assertIn("Special Form / all-risk", digest)
        self.assertIn("SAI-CPP-2024-07831", digest)
        self.assertIn("Claim Category Coverage Schedule", digest)
        self.assertIn("Specific Line-Item Determinations", digest)
        self.assertIn("CNC plasma cutting table #3", digest)
        self.assertIn("$485,000", digest)
        self.assertIn("raw steel", digest)
        self.assertIn("$145,000", digest)
        self.assertIn("Coverage Arithmetic", digest)
        self.assertIn("$890,000 / 47", digest)
        self.assertIn("$833,191", digest)
        self.assertIn("$18,500", digest)
        self.assertIn("if an excluded cause of loss results in a covered cause of loss", digest)
        self.assertIn("$112,500", digest)
        prompt = build_synthesis_prompt(state)
        self.assertIn("For insurance coverage-determination memoranda", prompt)
        self.assertIn("claim-category/line-item coverage schedule", prompt)

    def test_insurance_claim_comparison_digest_preserves_cpp_causation_and_bi_math(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="insurance/compare-insurance",
            question="Prepare a coverage analysis memo comparing Ridgeline's proof of loss against the insurance policy.",
            answer_schema={"deliverables": ["coverage-analysis-memo.docx"]},
            metadata={"practice_area": "insurance"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "policy-cpp-2024-08817.docx", "path": "policy-cpp-2024-08817.docx"},
                {"doc_id": "doc_2", "filename": "proof-of-loss-submission.docx", "path": "proof-of-loss-submission.docx"},
                {"doc_id": "doc_3", "filename": "maintenance-records-cooling-system.docx", "path": "maintenance-records-cooling-system.docx"},
                {"doc_id": "doc_4", "filename": "kellner-environmental-report.docx", "path": "kellner-environmental-report.docx"},
                {"doc_id": "doc_5", "filename": "bridgepoint-bi-report.docx", "path": "bridgepoint-bi-report.docx"},
                {"doc_id": "doc_6", "filename": "broker-correspondence.eml", "path": "broker-correspondence.eml"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Commercial Property Policy No. CPP-2024-08817 issued by Aldersgate Mutual Insurance Company "
                        "and Crestview Mutual Insurance Company to Ridgeline Manufacturing. Policy has Exclusion F for "
                        "corrosion, Provision 3 anti-concurrent causation, Exclusion J pollution exclusion with Exception 2 "
                        "for hostile-fire-caused pollution, a $350,000 pollution remediation sub-limit, a 72-hour Business "
                        "Income waiting period, EB-400 Equipment Breakdown endorsement with corrosion exclusion and separate "
                        "$25,000 deductible, a fire protective systems exception for sprinkler discharge, Ordinance or Law "
                        "coverage with a $500,000 sublimit, and replacement cost terms subject to actual cash value holdback."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c2",
                    "index": 1,
                    "text": (
                        "Ridgeline's proof of loss totals $4,730,000: building structural damage $1,420,000; contents / "
                        "machinery damage $1,560,000; environmental remediation costs $620,000; business interruption "
                        "$830,000; extra expense $185,000; ordinance or law compliance costs $115,000. The ordinance line "
                        "includes $85,000 for updated fire suppression and $30,000 for ADA-compliant egress modifications. "
                        "The fire damaged two Haas VF-6SS CNC machines and one Schuler hydraulic press sustained sprinkler damage."
                    ),
                },
                {
                    "doc_id": "doc_3",
                    "chunk_id": "c3",
                    "index": 2,
                    "text": (
                        "Cooling system maintenance records dated September 5, 2023 noted moderate surface oxidation on "
                        "brass fittings and recommended replacement. The brass fitting later failed from corrosion, causing "
                        "coolant spray, fire, sprinkler activation, and contamination."
                    ),
                },
                {
                    "doc_id": "doc_4",
                    "chunk_id": "c4",
                    "index": 3,
                    "text": (
                        "Kellner Environmental reported that the fire ruptured three 55-gallon drums of Solvent 142, causing "
                        "production-floor and subsurface contamination. The remediation budget includes a $110,000 groundwater "
                        "monitoring program. Ridgeline retained Kellner on January 18, 2025, Kellner began emergency containment "
                        "on January 20, 2025, and IDEM issued Administrative Compliance Order ENV-2025-0042 on January 22, 2025."
                    ),
                },
                {
                    "doc_id": "doc_5",
                    "chunk_id": "c5",
                    "index": 4,
                    "text": (
                        "The Bridgepoint business interruption report starts the restoration period on January 14, 2025. "
                        "Daily gross revenue impact is approximately $14,400. The BI claim is $830,000. The report treats "
                        "$350,000 in continuing expenses saved while separately listing $1,375,000 in idle production worker "
                        "wages as retained production wages."
                    ),
                },
                {
                    "doc_id": "doc_6",
                    "chunk_id": "c6",
                    "index": 5,
                    "text": (
                        "Broker correspondence discusses Indiana efficient proximate cause arguments, Ridgeline's insured "
                        "counterargument to anti-concurrent causation, Sue and Labor Condition 8, and whether remediation "
                        "begun before insurer consent can be separated into emergency mitigation and broader cleanup."
                    ),
                },
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("High-Priority Insurance Claim Comparison Matrix", digest)
        self.assertIn("Coverage Amount and Recommendation Schedule", digest)
        self.assertIn("Causation and Exclusion Analysis", digest)
        self.assertIn("CPP-2024-08817", digest)
        self.assertIn("Aldersgate Mutual", digest)
        self.assertIn("Crestview Mutual", digest)
        self.assertIn("Ridgeline", digest)
        self.assertIn("$4,730,000", digest)
        self.assertIn("$1,420,000", digest)
        self.assertIn("$1,560,000", digest)
        self.assertIn("$620,000", digest)
        self.assertIn("$830,000", digest)
        self.assertIn("$185,000", digest)
        self.assertIn("$115,000", digest)
        self.assertIn("corrosion of a brass fitting", digest)
        self.assertIn("September 5, 2023", digest)
        self.assertIn("moderate surface oxidation", digest)
        self.assertIn("Exclusion F", digest)
        self.assertIn("anti-concurrent causation", digest)
        self.assertIn("efficient proximate cause", digest)
        self.assertIn("Indiana", digest)
        self.assertIn("insured's counterargument", digest)
        self.assertIn("Exclusion J", digest)
        self.assertIn("Exception 2", digest)
        self.assertIn("$350,000 pollution-remediation sub-limit", digest)
        self.assertIn("$110,000 groundwater monitoring", digest)
        self.assertIn("72-hour waiting period", digest)
        self.assertIn("January 17, 2025", digest)
        self.assertIn("$43,200", digest)
        self.assertIn("$786,800", digest)
        self.assertIn("$350,000 as continuing expenses saved", digest)
        self.assertIn("$1,375,000", digest)
        self.assertIn("EB-400", digest)
        self.assertIn("$25,000 deductible", digest)
        self.assertIn("Schuler hydraulic press", digest)
        self.assertIn("sprinkler", digest)
        self.assertIn("$30,000 ADA", digest)
        self.assertIn("January 20, 2025", digest)
        self.assertIn("January 22, 2025", digest)
        self.assertIn("Sue and Labor", digest)
        self.assertIn("replacement cost versus actual cash value", digest)
        self.assertIn("$390,000 each", digest)
        self.assertIn("$485,000 each", digest)
        self.assertIn("coolant spray, fire, sprinkler response, and Solvent 142 contamination", digest)
        self.assertNotIn("SAI-CPP-2024-07831", digest)
        prompt = build_synthesis_prompt(state)
        self.assertIn("For CPP/Aldersgate insurance claim-comparison memoranda", prompt)
        self.assertIn("Do not reuse Sentinel", prompt)

    def test_insurance_policy_spec_digest_preserves_gap_matrix_and_acquisition_math(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="insurance/compare-commercial-insurance-policy-terms-against-coverage-specifications",
            question="Review the attached coverage specifications against the six issued policies and broker transmittal.",
            answer_schema={"deliverables": ["coverage-gap-analysis-memo.docx"]},
            metadata={"practice_area": "insurance"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "coverage-specifications.docx", "path": "coverage-specifications.docx"},
                {"doc_id": "doc_2", "filename": "cgl-policy-northland.docx", "path": "cgl-policy-northland.docx"},
                {"doc_id": "doc_3", "filename": "property-policy-northland.docx", "path": "property-policy-northland.docx"},
                {"doc_id": "doc_4", "filename": "excess-policy-atlantic.docx", "path": "excess-policy-atlantic.docx"},
                {"doc_id": "doc_5", "filename": "do-policy-commonwealth.docx", "path": "do-policy-commonwealth.docx"},
                {"doc_id": "doc_6", "filename": "epl-policy-commonwealth.docx", "path": "epl-policy-commonwealth.docx"},
                {"doc_id": "doc_7", "filename": "cyber-policy-ironshore.docx", "path": "cyber-policy-ironshore.docx"},
                {"doc_id": "doc_8", "filename": "broker-email-policies.eml", "path": "broker-email-policies.eml"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Coverage Specifications requested $2,000,000 products recall expense coverage, no implantable "
                        "medical device exclusion, CGL automatic acquisition coverage for 180 days, property newly acquired "
                        "locations $25,000,000 for 180 days, 18 months Business Income, $15,000,000 flood, excess defense "
                        "costs outside limits, no laser endorsement for Plant 4 defense operations, D&O automatic subsidiary "
                        "30% of Ridgeline assets, EPL third-party coverage, cyber social engineering $1,000,000, and a "
                        "$2,000,000 cyber bodily injury carve-back. Vantage has a July 15, 2025 closing, $38.6 million "
                        "facility replacement cost, $94 million assets, 385 employees, and $72 million revenue."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c2",
                    "index": 1,
                    "text": (
                        "Northland CGL has Recall of Products exclusion, products-completed aggregate $10,000,000, "
                        "newly acquired entities notice within 90 days, and NM-CGL-MDE-007 Implantable Medical Device "
                        "Exclusion. Defense costs do not reduce CGL limits and blanket additional insured and waiver of "
                        "subrogation are included."
                    ),
                },
                {
                    "doc_id": "doc_3",
                    "chunk_id": "c3",
                    "index": 2,
                    "text": (
                        "Northland Property provides a $175,000,000 blanket limit, replacement cost, agreed value, "
                        "12-month Period of Indemnity, $5,000,000 flood sublimit, and $10,000,000 newly acquired location "
                        "sublimit for 180 days."
                    ),
                },
                {
                    "doc_id": "doc_4",
                    "chunk_id": "c4",
                    "index": 3,
                    "text": (
                        "Atlantic Excess defense costs are included within and erode the $25,000,000 limits. Schedule A "
                        "lists CGL and Auto only, omitting Employers Liability. Plant 4 defense products are subject to a "
                        "$5,000,000 laser endorsement. Automatic acquisition coverage requires 90-day notice."
                    ),
                },
                {
                    "doc_id": "doc_5",
                    "chunk_id": "c5",
                    "index": 4,
                    "text": (
                        "Commonwealth D&O premium is $178,000. Optional ERP is 200% or $356,000. Automatic subsidiary "
                        "coverage applies only below fifteen percent (15%) of $245,000,000 assets."
                    ),
                },
                {
                    "doc_id": "doc_6",
                    "chunk_id": "c6",
                    "index": 5,
                    "text": "Commonwealth EPL includes a Third-Party Claims Exclusion and excludes claims by customers and vendors.",
                },
                {
                    "doc_id": "doc_7",
                    "chunk_id": "c7",
                    "index": 6,
                    "text": (
                        "Ironshore Cyber social engineering fraud sublimit is $250,000. Bodily Injury or Property Damage "
                        "exclusion applies without exception and no carve-back is included."
                    ),
                },
                {
                    "doc_id": "doc_8",
                    "chunk_id": "c8",
                    "index": 7,
                    "text": (
                        "Broker Diane Pressler stated coverages were placed in accordance with the specifications with "
                        "only minor adjustments."
                    ),
                },
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("High-Priority Insurance Policy Specification Gap Matrix", digest)
        self.assertIn("Near-Top Summary Table of Gaps", digest)
        self.assertIn("Vantage Acquisition Impact Calculations", digest)
        self.assertIn("Confirmed Matching Coverage Areas", digest)
        self.assertIn("Broker Representation / Follow-Up Checklist", digest)
        self.assertIn("ISSUE_001", digest)
        self.assertIn("products recall", digest)
        self.assertIn("$2,000,000", digest)
        self.assertIn("ISSUE_002", digest)
        self.assertIn("$25,000,000", digest)
        self.assertIn("$10,000,000", digest)
        self.assertIn("$38.6M", digest)
        self.assertIn("$28.6M", digest)
        self.assertIn("Critical", digest)
        self.assertIn("ISSUE_003", digest)
        self.assertIn("18-month", digest)
        self.assertIn("twelve (12) months", digest)
        self.assertIn("ISSUE_004", digest)
        self.assertIn("defense costs inside", digest)
        self.assertIn("ISSUE_005", digest)
        self.assertIn("$36.75M", digest)
        self.assertIn("$73.5M", digest)
        self.assertIn("Vantage assets $94M", digest)
        self.assertIn("ISSUE_006", digest)
        self.assertIn("third-party EPL", digest)
        self.assertIn("ISSUE_007", digest)
        self.assertIn("$750,000", digest)
        self.assertIn("ISSUE_008", digest)
        self.assertIn("October 13, 2025", digest)
        self.assertIn("ISSUE_009", digest)
        self.assertIn("$10,000,000 versus requested", digest)
        self.assertIn("ISSUE_010", digest)
        self.assertIn("bodily-injury carve-back", digest)
        self.assertIn("ISSUE_011", digest)
        self.assertIn("$59.3M / 19.0%", digest)
        self.assertIn("removal of the Plant 4 defense products laser endorsement", digest)
        self.assertIn("ISSUE_012", digest)
        self.assertIn("$178,000", digest)
        self.assertIn("ISSUE_013", digest)
        self.assertIn("implantable medical device", digest)
        self.assertIn("$71.5M / 22.9%", digest)
        self.assertIn("ISSUE_014", digest)
        self.assertIn("Add Employers Liability to the excess underlying schedule", digest)
        self.assertIn("Northland Mutual Insurance Company", digest)
        self.assertIn("Atlantic Specialty Underwriters", digest)
        self.assertIn("Ironshore Cyber Insurance Company", digest)
        self.assertIn("in accordance with specifications", digest)
        self.assertNotIn("SAI-CPP-2024-07831", digest)
        self.assertNotIn("CPP-2024-08817", digest)
        prompt = build_synthesis_prompt(state)
        self.assertIn("For commercial insurance policy/specification gap-analysis memoranda", prompt)
        self.assertIn("Put the gap summary table near the top", prompt)

    def test_environmental_indemnity_digest_preserves_lender_and_cost_issues(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="environmental-esg/analyze-counterparty-markup-of-environmental-indemnity-agreement",
            question="Analyze seller's redline of the environmental indemnity agreement.",
            answer_schema={"deliverables": ["eia-redline-analysis-memo.docx"]},
            metadata={"practice_area": "environmental-esg"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "buyers-draft-eia.docx", "extension": ".docx"},
                {"doc_id": "doc_2", "filename": "lender-term-sheet-enviro.docx", "extension": ".docx"},
                {"doc_id": "doc_3", "filename": "remediation-cost-estimate.xlsx", "extension": ".xlsx"},
                {"doc_id": "doc_4", "filename": "sellers-markup-eia.docx", "extension": ".docx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Buyer draft names Seller and Petrochem Legacy Management Inc. as GP, jointly and severally. "
                        "The draft has no cap and covers known or unknown Pre-Closing Environmental Conditions."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c2",
                    "index": 0,
                    "text": (
                        "Lender Section 7.1 says joint and several liability of the limited partnership and GP is a "
                        "fundamental credit underwriting requirement. Lender also requires $8,000,000 LOC, "
                        "$10,000,000/$20,000,000 PLL, residential direct contact standards, and freely assignable rights."
                    ),
                },
                {
                    "doc_id": "doc_3",
                    "chunk_id": "c3",
                    "index": 0,
                    "text": (
                        "Reasonable worst-case remediation is $14,406,000, leaving $594,000 under a $15,000,000 cap. "
                        "Vapor intrusion could cost $600,000 to $2,400,000 and residential versus industrial differential is about $1,200,000."
                    ),
                },
                {
                    "doc_id": "doc_4",
                    "chunk_id": "c4",
                    "index": 0,
                    "text": (
                        "Seller markup removes the GP, narrows coverage to specifically identified Phase II conditions, "
                        "uses Texas law and Houston arbitration, adds a surety bond, and reduces the LOC and PLL."
                    ),
                },
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("High-Priority Environmental Indemnity Matrix", digest)
        self.assertIn("GP deletion / credit support", digest)
        self.assertIn("known-or-unknown", digest)
        self.assertIn("$14,406,000", digest)
        self.assertIn("$594,000", digest)
        self.assertIn("$8,000,000", digest)
        self.assertIn("below five (5) feet", digest)
        self.assertIn("commercial-unavailability", digest)
        self.assertIn("$15,000,000 per occurrence / $25,000,000 aggregate", digest)
        self.assertIn("$10,000,000 per occurrence / $20,000,000 aggregate", digest)
        self.assertIn("Closing-Failure / Must-Have List", digest)
        self.assertIn("Rubric Preservation Checklist", digest)
        self.assertIn("Restore residential/unrestricted-use remediation standard", digest)
        prompt = build_synthesis_prompt(state)
        self.assertIn("For environmental indemnity redline memoranda", prompt)
        self.assertIn("High-Priority Environmental Indemnity Matrix", prompt)

    def test_environmental_esg_digest_routes_asaoc_and_settlement_tasks(self) -> None:
        asaoc_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="environmental-esg/analyze-counterparty-markup-of-administrative-settlement-agreement",
            question="Analyze Lakeview redline of the administrative settlement agreement.",
            answer_schema={"deliverables": ["redline-analysis-memorandum.docx"]},
            metadata={"practice_area": "environmental-esg"},
        )
        asaoc_state = RunState(task=asaoc_task, config=load_config(), documents=[])
        asaoc_contract = build_deliverable_contract(asaoc_state)
        asaoc_digest = build_task_family_digest(asaoc_state)
        self.assertEqual(asaoc_contract["task_family"], "environmental_esg_review")
        self.assertIn(
            "Near-top environmental / ESG required findings",
            asaoc_contract["deliverables"][0]["required_sections"],
        )
        self.assertIn("ASAOC / Administrative Settlement Review Rows", asaoc_digest)
        self.assertIn("40 C.F.R. 264.143(f)", asaoc_digest)
        self.assertIn("covenant not to sue", asaoc_digest)
        self.assertIn("250,000", asaoc_digest)
        self.assertIn("de novo", asaoc_digest)
        self.assertIn("30-day deemed-acceptance", asaoc_digest)
        self.assertIn("54 months to 78 months", asaoc_digest)
        self.assertIn("5,280,000 dollars to 4,730,000 dollars", asaoc_digest)

        settlement_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="environmental-esg/analyze-counterparty-markup-of-settlement-agreement",
            question="Analyze Saxonbrook redline of the environmental settlement agreement.",
            answer_schema={"deliverables": ["redline-review-memo.docx"]},
            metadata={"practice_area": "environmental-esg"},
        )
        settlement_state = RunState(task=settlement_task, config=load_config(), documents=[])
        settlement_digest = build_task_family_digest(settlement_state)
        self.assertIn("Environmental Settlement Agreement Rows", settlement_digest)
        self.assertIn("3,562,500", settlement_digest)
        self.assertIn("10,687,500", settlement_digest)
        self.assertIn("CERCLA Section 122(f)(6)", settlement_digest)
        self.assertIn("29,450,000", settlement_digest)
        self.assertIn("pre-2003 contamination", settlement_digest)
        self.assertIn("Oregon state law", settlement_digest)
        self.assertIn("1,500 / 3,000 / 5,000", settlement_digest)

    def test_environmental_esg_digest_routes_recall_and_esg_tasks(self) -> None:
        recall_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="environmental-esg/assess-recall-and-reporting-obligations-for-product-safety-issue",
            question="Draft an incident response memorandum for a product safety issue.",
            answer_schema={"deliverables": ["incident-response-memorandum.docx"]},
            metadata={"practice_area": "environmental-esg"},
        )
        recall_state = RunState(task=recall_task, config=load_config(), documents=[])
        recall_digest = build_task_family_digest(recall_state)
        self.assertIn("Product Safety Reporting and Recall Timeline", recall_digest)
        self.assertIn("CPSA Section 15(b)", recall_digest)
        self.assertIn("24 hours", recall_digest)
        self.assertIn("17.15 million", recall_digest)
        self.assertIn("28,000,000", recall_digest)
        self.assertIn("742,000", recall_digest)
        self.assertIn("257,000 pre-change", recall_digest)
        self.assertIn("zero fatalities", recall_digest)

        esg_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="environmental-esg/compare-esg-disclosure-against-regulatory-requirements",
            question="Review draft ESG report against regulatory requirements.",
            answer_schema={"deliverables": ["esg-gap-analysis-memo.docx"]},
            metadata={"practice_area": "environmental-esg"},
        )
        esg_state = RunState(task=esg_task, config=load_config(), documents=[])
        esg_contract = build_deliverable_contract(esg_state)
        esg_digest = build_task_family_digest(esg_state)
        self.assertEqual(esg_contract["task_family"], "environmental_esg_review")
        self.assertIn("ESG Disclosure Framework Gap Matrix", esg_digest)
        self.assertIn("2045", esg_digest)
        self.assertIn("3,840,000", esg_digest)
        self.assertIn("SB 253", esg_digest)
        self.assertIn("ESRS 2 GOV-3", esg_digest)
        self.assertIn("double materiality", esg_digest)
        self.assertIn("289,000 mtCO2e", esg_digest)
        self.assertIn("5 of 15 Scope 3 categories", esg_digest)
        self.assertIn("Poland / Gdansk", esg_digest)
        self.assertIn("1.5C, 2C", esg_digest)
        prompt = build_synthesis_prompt(esg_state)
        self.assertIn("For environmental, ESG, product-safety", prompt)
        self.assertIn("ESG Disclosure Framework Gap Matrix", prompt)

    def test_ip_contract_amendment_digest_preserves_provision_deltas_and_cover_omissions(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="intellectual-property/analyze-counterparty-markup-of-contract-amendment",
            question=(
                "Compare Veridian's redline against Pinnacle's draft amendment, cross-referencing the "
                "executed MSA, contracting policies, and correspondence."
            ),
            answer_schema={"deliverables": ["redline-deviation-report.docx"]},
            metadata={"practice_area": "intellectual-property"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "pinnacle-contracting-policy-tech.docx", "extension": ".docx"},
                {"doc_id": "doc_2", "filename": "pinnacle-internal-emails.eml", "extension": ".eml"},
                {"doc_id": "doc_3", "filename": "veridian-counsel-cover-email.eml", "extension": ".eml"},
                {"doc_id": "doc_4", "filename": "veridian-markup-amendment-redline.docx", "extension": ".docx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Pinnacle policy requires minimum 1.5x Annual Fees and prefers 2.0x. "
                        "At $17.47 million Annual Fees, the minimum cap is $26.205 million and "
                        "the preferred cap is $34.94 million. HIPAA/data security must be carved out."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c2",
                    "index": 1,
                    "text": (
                        "Internal email says the PHM Module must use 99.95% uptime, policy floor 99.9%, "
                        "2% service credits per 0.01% shortfall capped at 15%, and breach notice within "
                        "24 hours because North Carolina law requires notice as expeditiously as possible."
                    ),
                },
                {
                    "doc_id": "doc_3",
                    "chunk_id": "c3",
                    "index": 2,
                    "text": (
                        "Veridian cover email calls the CPI floor a minor 2.0% adjustment and describes "
                        "other changes as market standard and conforming edits without highlighting change "
                        "of control, governing law, audit restrictions, or HIPAA carve-out deletion."
                    ),
                },
                {
                    "doc_id": "doc_4",
                    "chunk_id": "c4",
                    "index": 3,
                    "text": (
                        "Veridian markup changes the cap to 1x, adds a data security incidents consequential "
                        "damages exclusion, sets PHM SLA at 99.5% with 1% credits and 5% cap, changes notice "
                        "to 365 days, ETF to 75%, auto-renewal to one 3-year term with 270 days notice, "
                        "transition to 6 months at 150%, permits PHM subcontractors without prior written "
                        "consent, changes breach notice to 30 days, limits audits to once yearly on 60 "
                        "business days notice, excludes subcontractor facilities, shifts audit costs above "
                        "$25,000, and changes North Carolina/Mecklenburg to Texas/Dallas."
                    ),
                },
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("High-Priority Technology Contract Amendment Deviation Matrix", digest)
        self.assertIn("Financial Exposure Calculations", digest)
        self.assertIn("Cover Email Omission Checklist", digest)
        self.assertIn("1.5x", digest)
        self.assertIn("$26.205M", digest)
        self.assertIn("$34.94M", digest)
        self.assertIn("HIPAA/data-security liability carve-out deletion", digest)
        self.assertIn("data-breach consequential-damages exclusion", digest)
        self.assertIn("99.5%", digest)
        self.assertIn("99.9%", digest)
        self.assertIn("1% credit rate", digest)
        self.assertIn("5% cap", digest)
        self.assertIn("$349,400", digest)
        self.assertIn("365-day", digest)
        self.assertIn("$13.1025M", digest)
        self.assertIn("$52.41M", digest)
        self.assertIn("270-day", digest)
        self.assertIn("6-month/150%", digest)
        self.assertIn("30 business days", digest)
        self.assertIn("without Pinnacle's prior written consent", digest)
        self.assertIn("30 calendar days", digest)
        self.assertIn("60 business days", digest)
        self.assertIn("$25,000", digest)
        self.assertIn("North Carolina to Texas", digest)
        self.assertIn("Change of control", digest)
        self.assertIn("audit restrictions", digest)
        self.assertIn("Potentially acceptable / immaterial edits", digest)
        prompt = build_synthesis_prompt(state)
        self.assertIn("For healthcare technology contract-amendment deviation reports", prompt)
        self.assertIn("High-Priority Technology Contract Amendment Deviation Matrix", prompt)

    def test_technology_data_agreement_digest_preserves_msa_data_clause_deltas(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="intellectual-property/analyze-counterparty-markup-of-master-services-agreement",
            question=(
                "Analyze Nexora's counterparty markup of the master services agreement against Verdantis's "
                "template, contract playbook, internal memo, SOW, and counsel cover email."
            ),
            answer_schema={"deliverables": ["msa-redline-review.docx"]},
            metadata={"practice_area": "intellectual-property"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "nexora-redline-msa.docx", "extension": ".docx"},
                {"doc_id": "doc_2", "filename": "verdantis-contracts-playbook.docx", "extension": ".docx"},
                {"doc_id": "doc_3", "filename": "draft-sow-001.docx", "extension": ".docx"},
                {"doc_id": "doc_4", "filename": "verdantis-internal-memo.docx", "extension": ".docx"},
                {"doc_id": "doc_5", "filename": "ashford-merritt-cover-email.eml", "extension": ".eml"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Section 9.1 changes the cap from TWO (2) TIMES THE TOTAL FEES PAID OR PAYABLE "
                        "during the TWELVE (12)-MONTH PERIOD to ONE (1) TIMES THE FEES ACTUALLY PAID "
                        "during the SIX (6)-MONTH period. Section 8.1(b) changes data breach indemnity "
                        "from negligence to gross negligence or willful misconduct and limits recovery to direct damages only."
                    ),
                },
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c2",
                    "index": 1,
                    "text": (
                        "Section 7.1(d) gives Vendor ownership of models, model weights, and improvements "
                        "developed using Customer Data, provided no Customer Data or derivatives are included. "
                        "Section 6.2 says the Business Associate Agreement may be signed within sixty (60) days. "
                        "Section 13 moves disputes to Western Arbitration Council arbitration and bars punitive or exemplary damages."
                    ),
                },
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c3",
                    "index": 2,
                    "text": (
                        "Section 10.2 requires a seventy-five percent (75%) early termination fee and Section 10.3 "
                        "extends cure from 30 to forty-five (45) days with additional time as reasonably necessary. "
                        "Section 14.3 deletes the M&A assignment exception."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c4",
                    "index": 3,
                    "text": (
                        "Verdantis playbook treats offshore processing, Singapore access, BAA timing, and data "
                        "residency as healthcare data escalation items. North Carolina contract claims normally "
                        "use N.C. Gen. Stat. Section 1-52(1)."
                    ),
                },
                {
                    "doc_id": "doc_3",
                    "chunk_id": "c5",
                    "index": 4,
                    "text": (
                        "Year 1 Platform Fee is $1,450,000; quarterly invoices are $362,500 and monthly platform "
                        "fees are $120,833.33. Year 2 is $1,493,500 and Year 3 is $1,538,305. SLA credits are "
                        "capped at five percent (5%) annually."
                    ),
                },
                {
                    "doc_id": "doc_4",
                    "chunk_id": "c6",
                    "index": 5,
                    "text": (
                        "Total Platform Fees are $4,481,805, the implementation fee is $385,000, and grand total "
                        "contract value is $4,866,805. Current security assessment flagged Singapore offshore "
                        "development for the healthcare platform."
                    ),
                },
                {
                    "doc_id": "doc_5",
                    "chunk_id": "c7",
                    "index": 6,
                    "text": (
                        "Cover email says the machine learning ownership language is standard hygiene, offshore "
                        "flexibility is operational, BAA timing within 60 days is practical, and California arbitration "
                        "is a standard commercial term."
                    ),
                },
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("High-Priority Technology/Data Agreement Clause-Delta Matrix", digest)
        self.assertIn("Data/Privacy/Security Control Checklist", digest)
        self.assertIn("Financial Exposure Calculations", digest)
        self.assertIn("Cover Email Omission Checklist", digest)
        self.assertIn("two times (2x)", digest)
        self.assertIn("one time (1x)", digest)
        self.assertIn("twelve (12)-month", digest)
        self.assertIn("six (6)-month", digest)
        self.assertIn("paid or payable", digest)
        self.assertIn("actually paid", digest)
        self.assertIn("Critical / highest-severity issue", digest)
        self.assertIn("$362,500", digest)
        self.assertIn("$725,000", digest)
        self.assertIn("gross negligence", digest)
        self.assertIn("$3,000,000 aggregate cap", digest)
        self.assertIn("direct damages only", digest)
        self.assertIn("Customer Data or derivatives", digest)
        self.assertIn("offshore", digest)
        self.assertIn("Singapore", digest)
        self.assertIn("within sixty (60) days", digest)
        self.assertIn("$2,273,854", digest)
        self.assertIn("~97%", digest)
        self.assertIn("12-month contractual limitations period", digest)
        self.assertIn("N.C. Gen. Stat. Section 1-52(1)", digest)
        self.assertIn("Western Arbitration Council", digest)
        self.assertIn("punitive or exemplary damages", digest)
        self.assertIn("$72,500", digest)
        self.assertIn("45 days", digest)
        self.assertIn("M&A assignment exception", digest)
        self.assertIn("1.5% per month", digest)
        self.assertIn("18% annualized", digest)
        self.assertIn("N.C. Gen. Stat. Section 24-1", digest)
        self.assertIn("$10,000,000 to $5,000,000", digest)
        self.assertIn("five (5) years to three (3) years", digest)
        self.assertIn("indefinite to five (5) years", digest)
        self.assertIn("60 days to 120 days", digest)
        self.assertIn("automatic SOW term extension", digest)
        self.assertIn("adequate security measures to commercially reasonable security measures", digest)
        self.assertIn("cap remains at 2x annual fees", digest)
        prompt = build_synthesis_prompt(state)
        self.assertIn("For technology, SaaS, MSA", prompt)
        self.assertIn("High-Priority Technology/Data Agreement Clause-Delta Matrix", prompt)

    def test_technology_data_agreement_digest_preserves_saas_and_dpa_controls(self) -> None:
        saas_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="intellectual-property/analyze-counterparty-markup-of-saas-agreement",
            question="Analyze the Cumulus/NovaSphere SaaS redline against Thorngate's template and order form.",
            answer_schema={"deliverables": ["saas-redline-review.docx"]},
            metadata={"practice_area": "intellectual-property"},
        )
        saas_state = RunState(
            task=saas_task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "cumulus-redline-markup.docx", "extension": ".docx"},
                {"doc_id": "doc_2", "filename": "thorngate-saas-template-clean.docx", "extension": ".docx"},
                {"doc_id": "doc_3", "filename": "order-form-sow-executed.docx", "extension": ".docx"},
                {"doc_id": "doc_4", "filename": "cio-deal-priorities.eml", "extension": ".eml"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Cumulus SaaS markup changes uptime from 99.9% to 99.5%, caps service credits at 10%, "
                        "extends data return from 30 to 90 days, extends deletion from 60 to 180 days, adds AS IS "
                        "third-party integration disclaimers, and changes governing law from Ohio to Texas."
                    ),
                },
                {
                    "doc_id": "doc_4",
                    "chunk_id": "c2",
                    "index": 1,
                    "text": (
                        "Thorngate is public and cares about SOX 404. The board-approved budget cap is $10.5M. "
                        "Ferriston Industrial Group may acquire Thorngate. Cumulus had an August 2023 data breach "
                        "and is a VC-backed Series D company."
                    ),
                },
            ],
        )
        saas_digest = build_task_family_digest(saas_state)
        self.assertIn("$3,700,000 to $1,850,000", saas_digest)
        self.assertIn("August 2023 Cumulus/NovaSphere data breach", saas_digest)
        self.assertIn("99.9% to 99.5%", saas_digest)
        self.assertIn("15% to 10%", saas_digest)
        self.assertIn("right to terminate for chronic SLA underperformance", saas_digest)
        self.assertIn("$10.5M", saas_digest)
        self.assertIn("30 to 90 days", saas_digest)
        self.assertIn("60 to 180 days", saas_digest)
        self.assertIn("VC-backed Series D", saas_digest)
        self.assertIn("$5,000,000 umbrella", saas_digest)
        self.assertIn("Ohio to Texas", saas_digest)
        self.assertIn("Summit County / N.D. Ohio to Travis County / W.D. Texas", saas_digest)
        self.assertIn("Ferriston Industrial Group", saas_digest)
        self.assertIn("vendor assignment to an affiliate", saas_digest)
        self.assertIn("SOC 2 Type II report and ISO 27001 certificate only", saas_digest)
        self.assertIn("SOX 404", saas_digest)
        self.assertIn("standalone data breach indemnification", saas_digest)
        self.assertIn("IP infringement remedies", saas_digest)

        dpa_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="data-privacy-cybersecurity/analyze-counterparty-markup-of-data-processing-agreement",
            question="Analyze CloudNest's redlined data processing agreement against Stratton's DPA playbook.",
            answer_schema={"deliverables": ["dpa-redline-review.docx"]},
            metadata={"practice_area": "data-privacy-cybersecurity"},
        )
        dpa_state = RunState(
            task=dpa_task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "cloudnest-redlined-dpa.docx", "extension": ".docx"},
                {"doc_id": "doc_2", "filename": "stratton-health-dpa-playbook.docx", "extension": ".docx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "CloudNest DPA reduces subprocessor notice from 30 to 15 days, adds Peregrine Data Analytics "
                        "in Mumbai, India for log analytics and performance monitoring, extends breach notice from "
                        "24 hours to 72 hours after confirming the incident, removes HITRUST CSF certification, deletes "
                        "$50M per occurrence / $100M aggregate cyber insurance, changes DSR support to 15 business days "
                        "with fees after 10 requests per month, uses England and Wales law with London courts, and "
                        "decouples the DPA with a 180-day notice period."
                    ),
                }
            ],
        )
        dpa_digest = build_task_family_digest(dpa_state)
        self.assertIn("72 hours", dpa_digest)
        self.assertNotIn("Notice timing is loosened from 24 hours to 48 hours", dpa_digest)
        self.assertIn("30 days to 15 days", dpa_digest)
        self.assertIn("right to object and terminate", dpa_digest)
        self.assertIn("log analytics and performance monitoring", dpa_digest)
        self.assertIn("36-hour Red threshold", dpa_digest)
        self.assertIn("30 business days exceeds the playbook's 20-day threshold", dpa_digest)
        self.assertIn("India lacks an EU adequacy decision", dpa_digest)
        self.assertIn("$18.6M", dpa_digest)
        self.assertIn("$55.8M", dpa_digest)
        self.assertIn("$37.2M", dpa_digest)
        self.assertIn("HITRUST CSF", dpa_digest)
        self.assertIn("$50M per occurrence / $100M aggregate", dpa_digest)
        self.assertIn("Classify the 60-day data-return change as Yellow", dpa_digest)
        self.assertIn("classify the 120-day deletion change as Red", dpa_digest)
        self.assertIn("upon reasonable request", dpa_digest)
        self.assertIn("England and Wales / London change as Red", dpa_digest)
        self.assertIn("impact on liability, indemnification, remedies", dpa_digest)
        self.assertIn("processor-only indemnification to mutual indemnification", dpa_digest)
        self.assertIn("2.3 million U.S. patients plus 14,000 EU/UK patients", dpa_digest)

    def test_prenuptial_digest_preserves_cross_provision_asset_rights(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="trusts-estates-private-client/analyze-counterparty-markup-of-prenuptial-agreement",
            question="Review the counterparty redline against our original prenuptial agreement.",
            answer_schema={"deliverables": ["redline-analysis-memo.docx"]},
            metadata={"practice_area": "trusts-estates-private-client"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "counterparty-redline-markup.docx", "path": "counterparty-redline-markup.docx"},
                {"doc_id": "doc_2", "filename": "financial-disclosure-summary.xlsx", "path": "financial-disclosure-summary.xlsx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Section 3.2 says Active Appreciation on Separate Property shall be classified as Marital Property. "
                        "Section 4.1(c) grants Marcus Brannigan fifteen percent (15%) of Net After-Tax Proceeds above the "
                        "$24,080,000 OrthoNova Baseline Value and says this is in addition to, and shall not limit or offset, "
                        "any other claims. Section 4.2 says appreciation in Brannigan Capital, passive or active, remains "
                        "Brannigan's Separate Property. OrthoNova may have an IPO in 18 to 36 months."
                    ),
                },
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c2",
                    "index": 1,
                    "text": (
                        "Section 5.3 reimburses Capital Contributions including mortgage payments made from Separate Income. "
                        "Section 6.1 creates a $250,000 separate income threshold. Section 7.3 allows a neutral financial planner "
                        "selected by the payee spouse. Section 9.4 preserves elective share rights; Section 9.2 only protects an "
                        "irrevocable trust for Sophia. Section 13.2 requires binding arbitration and strict confidentiality."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c3",
                    "index": 2,
                    "text": (
                        "Dr. Hartley-Chen certifies under penalty of perjury that her disclosure is complete and accurate. "
                        "Marcus Brannigan certifies only a substantially accurate summary. His disclosure lists 4 of 7 commercial "
                        "properties. Hartley-Chen earned $620,000; Brannigan earned $485,000."
                    ),
                },
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("High-Priority Prenuptial Asset-Rights Matrix", digest)
        self.assertIn("Section 3.2 / 4.2 active-appreciation asymmetry", digest)
        self.assertIn("Double-counting", digest)
        self.assertIn("$3,612,000", digest)
        self.assertIn("18-36 months", digest)
        self.assertIn("pre-marital goodwill", digest)
        self.assertIn("original was 25% of marriage duration capped at 36 months", digest)
        self.assertIn("61%", digest)
        self.assertIn("39%", digest)
        self.assertIn("neutral financial planner", digest)
        self.assertIn("Conn. Gen. Stat. Section 45a-436", digest)
        self.assertIn("substantially accurate summary", digest)
        self.assertIn("4 properties", digest)
        self.assertIn("cover letter omissions", digest.lower())
        prompt = build_synthesis_prompt(state)
        self.assertIn("For prenuptial, premarital", prompt)
        self.assertIn("High-Priority Prenuptial Asset-Rights Matrix", prompt)

    def test_trusts_estates_digest_routes_parenting_and_postnup_tasks(self) -> None:
        parenting_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="trusts-estates-private-client/analyze-counterparty-markup-of-parenting-plan",
            question="Analyze counterparty markup of parenting plan.",
            answer_schema={"deliverables": ["redline-analysis-memorandum.docx"]},
            metadata={"practice_area": "trusts-estates-private-client"},
        )
        parenting_state = RunState(task=parenting_task, config=load_config(), documents=[])
        parenting_contract = build_deliverable_contract(parenting_state)
        parenting_digest = build_task_family_digest(parenting_state)
        self.assertEqual(parenting_contract["task_family"], "trusts_estates_private_client_review")
        self.assertIn("Near-top trusts and estates required findings", parenting_contract["deliverables"][0]["required_sections"])
        self.assertIn("Rodrigo travels 6-8 days", parenting_digest)
        self.assertIn("60-to-14-day notice change", parenting_digest)
        self.assertIn("156 x 18.4 = 2,870", parenting_digest)
        self.assertIn("Sofia's soccer cost", parenting_digest)

        postnup_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="trusts-estates-private-client/analyze-counterparty-markup-of-postnuptial-agreement",
            question="Analyze counterparty markup of postnuptial agreement.",
            answer_schema={"deliverables": ["redline-analysis-memo.docx"]},
            metadata={"practice_area": "trusts-estates-private-client"},
        )
        postnup_digest = build_task_family_digest(RunState(task=postnup_task, config=load_config(), documents=[]))
        self.assertIn("Husband's disclosed inheritance is about $3.8M", postnup_digest)
        self.assertIn("$1.2M and $600K", postnup_digest)
        self.assertIn("12,000 x $38 = $456K", postnup_digest)
        self.assertIn("750 ILCS 5/502(b)", postnup_digest)

    def test_trusts_estates_digest_routes_charitable_and_estate_claim_tasks(self) -> None:
        charitable_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="trusts-estates-private-client/compare-charitable-trust-structures-against-client-goals",
            question="Compare charitable trust structures against client goals, including income tax effects.",
            answer_schema={"deliverables": ["trust-comparison-memo.docx"]},
            metadata={"practice_area": "trusts-estates-private-client"},
        )
        charitable_state = RunState(task=charitable_task, config=load_config(), documents=[])
        charitable_contract = build_deliverable_contract(charitable_state)
        charitable_digest = build_task_family_digest(charitable_state)
        self.assertEqual(charitable_contract["task_family"], "trusts_estates_private_client_review")
        self.assertIn("Near-top trusts and estates required findings", charitable_contract["deliverables"][0]["required_sections"])
        self.assertNotIn("Tax issue matrix", charitable_contract["deliverables"][0]["required_sections"])
        self.assertIn("$370K / $555K", charitable_digest)
        self.assertIn("$18.2M for the CRAT and $16.4M for the CRUT", charitable_digest)
        self.assertIn("$12.374M for the CRAT and $10.186M for the CRUT", charitable_digest)

        estate_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="trusts-estates-private-client/compare-creditor-claims-against-estate-assets",
            question="Compare creditor claims against estate assets and tax authority claims.",
            answer_schema={"deliverables": ["claims-reconciliation-memo.docx"]},
            metadata={"practice_area": "trusts-estates-private-client"},
        )
        estate_state = RunState(task=estate_task, config=load_config(), documents=[])
        estate_contract = build_deliverable_contract(estate_state)
        estate_digest = build_task_family_digest(estate_state)
        self.assertEqual(estate_contract["task_family"], "trusts_estates_private_client_review")
        self.assertIn("Total probate estate is about $4,631,870", estate_digest)
        self.assertIn("755 ILCS 5/18-10", estate_digest)
        self.assertIn("$500K life insurance or $347,200 IRA", estate_digest)
        self.assertIn("Natalie 40%, Daniel 35%, and Foundation 15%", estate_digest)

    def test_healthcare_digest_routes_compliance_and_cta_tasks(self) -> None:
        compliance_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="healthcare-life-sciences/analyze-compliance-program-gaps",
            question="Review HIPAA compliance program materials and prepare a gap analysis memo.",
            answer_schema={"deliverables": ["compliance-gap-memorandum.docx"]},
            metadata={"practice_area": "healthcare-life-sciences"},
        )
        compliance_state = RunState(task=compliance_task, config=load_config(), documents=[])
        compliance_contract = build_deliverable_contract(compliance_state)
        compliance_digest = build_task_family_digest(compliance_state)
        self.assertEqual(compliance_contract["task_family"], "healthcare_life_sciences_review")
        self.assertIn(
            "Near-top healthcare / life-sciences required findings",
            compliance_contract["deliverables"][0]["required_sections"],
        )
        self.assertIn("About 9 of 47 vendors", compliance_digest)
        self.assertIn("45 CFR 164.406", compliance_digest)
        self.assertIn("312 employees", compliance_digest)
        self.assertIn("164.522(a)(1)(vi)", compliance_digest)

        cta_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="healthcare-life-sciences/analyze-counterparty-markup-of-clinical-trial-agreement",
            question="Analyze counterparty markup of clinical trial agreement.",
            answer_schema={"deliverables": ["redline-analysis-memo.docx"]},
            metadata={"practice_area": "healthcare-life-sciences"},
        )
        cta_digest = build_task_family_digest(RunState(task=cta_task, config=load_config(), documents=[]))
        self.assertIn("Lakeshore nonprofit status", cta_digest)
        self.assertIn("7 years to 3 years", cta_digest)
        self.assertIn("12-month wind-down payment", cta_digest)
        self.assertIn("21 CFR 312.62(c)", cta_digest)

    def test_healthcare_digest_routes_merger_protocol_and_certificate_tasks(self) -> None:
        merger_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="healthcare-life-sciences/analyze-counterparty-markup-of-merger-agreement",
            question="Analyze counterparty markup of healthcare merger agreement.",
            answer_schema={"deliverables": ["redline-analysis-memorandum.docx"]},
            metadata={"practice_area": "healthcare-life-sciences"},
        )
        merger_digest = build_task_family_digest(RunState(task=merger_task, config=load_config(), documents=[]))
        self.assertIn("$156 floor, $201 cap", merger_digest)
        self.assertIn("$72.00 to $78.00", merger_digest)
        self.assertIn("will to should", merger_digest)

        protocol_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="healthcare-life-sciences/compare-clinical-trial-protocol-against-fda-regulatory-requirements",
            question="Compare clinical trial protocol against FDA regulatory requirements.",
            answer_schema={"deliverables": ["gap-analysis-memorandum.docx"]},
            metadata={"practice_area": "healthcare-life-sciences"},
        )
        protocol_digest = build_task_family_digest(RunState(task=protocol_task, config=load_config(), documents=[]))
        self.assertIn("21 CFR 50.25(a)(5)", protocol_digest)
        self.assertIn("two independent hepatopathologists", protocol_digest)
        self.assertIn("variceal bleeding", protocol_digest)

        certificate_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="healthcare-life-sciences/compare-closing-certificate-against-agreement-covenants",
            question="Compare closing certificate against merger agreement covenants.",
            answer_schema={"deliverables": ["compliance-gap-analysis-memo.docx"]},
            metadata={"practice_area": "healthcare-life-sciences"},
        )
        certificate_digest = build_task_family_digest(RunState(task=certificate_task, config=load_config(), documents=[]))
        self.assertIn("$1,475,000 exceeds the $1,250,000", certificate_digest)
        self.assertIn("Section 4(f) reports $1,800,000", certificate_digest)
        self.assertIn("Dr. Helen Vargas's salary increase from $285,000 to $310,000", certificate_digest)
        self.assertIn("$24,380,000 on March 31, 2025", certificate_digest)
        self.assertIn("Section 4(m) reports only the April 25, 2025 cash balance", certificate_digest)
        self.assertIn("338(h)(10)", certificate_digest)
        self.assertIn("$2,300,000, above the $750,000", certificate_digest)
        self.assertIn("Explicitly call the Apex CRO MSA a Material Contract", certificate_digest)
        self.assertIn("11-day delay", certificate_digest)
        self.assertIn("Section 6.2(l) for unauthorized hiring", certificate_digest)

    def test_corporate_ma_digest_preserves_change_of_control_contract_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="corporate-ma/analyze-change-of-control-provisions-across-targets-material-contracts",
            question="Analyze change of control provisions across the target's material contracts.",
            answer_schema={"deliverables": ["coc-analysis-report.docx"]},
            metadata={"practice_area": "corporate-ma"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "deal-overview-memo.docx", "extension": ".docx"},
                {"doc_id": "doc_2", "filename": "terranode-isa.docx", "extension": ".docx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Aldersgate uses TerraNode for all production cloud infrastructure. "
                        "CloudSpan is a Ridgeline portfolio company in cloud infrastructure. "
                        "Expected closing is October 15, 2025."
                    ),
                }
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("Near-Top Corporate M&A Required Findings", digest)
        self.assertIn("TerraNode direct-competitor consent risk", digest)
        self.assertIn("$6.2M", digest)
        self.assertIn("CloudSpan", digest)
        self.assertIn("30-day termination notice / 60-day cure", digest)
        self.assertIn("sole-discretion/no-leverage", digest)
        self.assertIn("September 15, 2025", digest)
        self.assertIn("$35.4076M", digest)
        self.assertIn("24 months base salary ($850K)", digest)
        self.assertIn("ultimate controlling person", digest)
        self.assertNotIn("High-Priority Venture Financing Issue Matrix", digest)
        prompt = build_synthesis_prompt(state)
        self.assertIn("For corporate M&A diligence", prompt)
        self.assertIn("Deal Math / Timeline / Exposure Calculations", prompt)

    def test_corporate_ma_digest_preserves_spa_markup_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="corporate-ma/analyze-counterparty-spa-markup",
            question="Analyze the Seller markup of the DataForge stock purchase agreement.",
            answer_schema={"deliverables": ["markup-deviation-report.docx"]},
            metadata={"practice_area": "corporate-ma"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "buyer-negotiation-playbook.docx", "extension": ".docx"},
                {"doc_id": "doc_2", "filename": "seller-markup-spa.docx", "extension": ".docx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": "DataForge Seller markup SPA adds earnout and deferred consideration.",
                }
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("Corporate SPA Markup Issue Matrix", digest)
        self.assertIn("$72M is about 12.1%", digest)
        self.assertIn("commercially reasonable efforts", digest)
        self.assertIn("45-position reduction", digest)
        self.assertIn("two $7.75M payments", digest)
        self.assertIn("Michael Huang", digest)
        self.assertIn("GPLv3", digest)
        self.assertIn("3% of equity value", digest)
        self.assertIn("August 31, 2025", digest)
        self.assertIn("minimum 15 months", digest)
        self.assertIn("Document Cover-Letter Omission Checklist", digest)

    def test_corporate_ma_digest_preserves_cim_and_credit_calculations(self) -> None:
        cim_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="corporate-ma/analyze-cim-deal-teaser/scenario-01",
            question="Analyze the CIM and deal teaser for Cascade Environmental.",
            answer_schema={"deliverables": ["cim-diligence-memo.docx"]},
            metadata={"practice_area": "corporate-ma"},
        )
        cim_state = RunState(
            task=cim_task,
            config=load_config(),
            documents=[{"doc_id": "doc_1", "filename": "ces-cim.docx", "extension": ".docx"}],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": "Cascade Environmental deal teaser reports EBITDA, DSO, PFAS, and top 10 customers.",
                }
            ],
        )
        cim_digest = build_task_family_digest(cim_state)
        self.assertIn("CIM / Deal-Teaser Diligence Issue Matrix", cim_digest)
        self.assertIn("$16.8M, not $17.1M", cim_digest)
        self.assertIn("About 70.6 days", cim_digest)
        self.assertIn("$29.4M", cim_digest)
        self.assertIn("firm-versus-soft backlog", cim_digest)
        self.assertIn("18-24 months post-close", cim_digest)
        self.assertIn("hazardous waste transporter permits", cim_digest)

        credit_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="corporate-ma/analyze-credit-agreement-markup",
            question="Analyze the borrower markup of the acquisition credit agreement.",
            answer_schema={"deliverables": ["credit-markup-analysis.docx"]},
            metadata={"practice_area": "corporate-ma"},
        )
        credit_state = RunState(
            task=credit_task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "arranger-analysis-template.xlsx", "extension": ".xlsx"},
                {"doc_id": "doc_2", "filename": "borrower-markup-v2.docx", "extension": ".docx"},
            ],
            chunks=[{"doc_id": "doc_2", "chunk_id": "c1", "index": 0, "text": "EverBright borrower markup changes equity cure and ECF sweep."}],
        )
        credit_digest = build_task_family_digest(credit_state)
        self.assertIn("Acquisition Financing Markup Issue Matrix", credit_digest)
        self.assertIn("20 business days", credit_digest)
        self.assertIn("no-over-cure", credit_digest)
        self.assertIn("$14.0M", credit_digest)
        self.assertIn("junior-lien incremental", credit_digest)
        self.assertIn("CLOs managed by DQ Lenders", credit_digest)
        self.assertIn("cash-netting-cap deletion", credit_digest)
        self.assertIn("consecutive quarter cures", credit_digest)
        self.assertIn("New York law is market standard", credit_digest)
        self.assertIn("J. Crew", credit_digest)
        self.assertIn("Available Equity Amount", credit_digest)
        self.assertIn("excludes all letters of credit", credit_digest)

    def test_energy_project_finance_digest_preserves_bankability_and_calculations(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="energy-natural-resources/analyze-counterparty-markup-of-power-purchase-agreement",
            question="Review the buyer markup of the power purchase agreement for project finance risks.",
            answer_schema={"deliverables": ["ppa-markup-analysis.docx"]},
            metadata={"practice_area": "energy-natural-resources"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "seller-playbook.docx", "extension": ".docx"},
                {"doc_id": "doc_2", "filename": "buyer-redline-power-purchase-agreement.docx", "extension": ".docx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Seller playbook requires at least a 60-day buffer between Target COD and Guaranteed COD, "
                        "uses 70% of 575,000 MWh as the performance termination threshold, and preserves Texas "
                        "Local Government Code Chapter 271 waiver language."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c2",
                    "index": 1,
                    "text": (
                        "Buyer markup moves Guaranteed COD to December 31, 2026, increases delay LDs to "
                        "$2,000/MW/day, raises MAEP to 90% of P50, deletes the sovereign-immunity waiver, "
                        "permits fund-level transfers by Solara Infrastructure Partners LP, and requires "
                        "575,000 RECs/year."
                    ),
                },
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("High-Priority Energy Project Finance Issue Matrix", digest)
        self.assertIn("Near-Top Energy Required Findings", digest)
        self.assertIn("Power Purchase Agreement Issues", digest)
        self.assertIn("lender requires at least 60-day buffer", digest)
        self.assertIn("460,000 MWh to 517,500 MWh", digest)
        self.assertIn("402,500 MWh vs 431,250 MWh", digest)
        self.assertIn("575,000 RECs/year", digest)
        self.assertIn("Chapter 271", digest)
        self.assertIn("Project Schedule / LD / Revenue Calculations", digest)
        self.assertNotIn("High-Priority Venture Financing Issue Matrix", digest)
        prompt = build_synthesis_prompt(state)
        self.assertIn("For energy, project-finance", prompt)
        self.assertIn("Near-Top Energy Required Findings", prompt)
        self.assertIn("Bankability and Lender-Control Checklist", prompt)

    def test_energy_project_finance_digest_routes_credit_epc_intercreditor_concession(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="energy-natural-resources/analyze-counterparty-markup-of-credit-agreement",
            question=(
                "Analyze the concession agreement, credit agreement, EPC contract, and intercreditor agreement "
                "for project finance markup issues."
            ),
            answer_schema={"deliverables": ["energy-markup-analysis.docx"]},
            metadata={"practice_area": "energy-natural-resources"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "ridgeline-credit-agreement-redline.docx", "extension": ".docx"},
                {"doc_id": "doc_2", "filename": "epc-contract-contractor-markup.docx", "extension": ".docx"},
                {"doc_id": "doc_3", "filename": "intercreditor-agreement-markup.docx", "extension": ".docx"},
                {"doc_id": "doc_4", "filename": "concession-agreement-authority-markup.docx", "extension": ".docx"},
            ],
            chunks=[],
        )
        digest = build_task_family_digest(state)
        self.assertIn("Concession Agreement Issues", digest)
        self.assertIn("50/50 and later 25% Concessionaire / 75% TxDOT", digest)
        self.assertIn("25% Concessionaire / 75% TxDOT", digest)
        self.assertIn("60% Concessionaire / 40% TxDOT", digest)
        self.assertIn("Energy Credit Agreement Issues", digest)
        self.assertIn("$5M/month", digest)
        self.assertIn("66 2/3%", digest)
        self.assertIn("EPC Contract Issues", digest)
        self.assertIn("$115,000 per day", digest)
        self.assertIn("November 30, 2026", digest)
        self.assertIn("Intercreditor Agreement Issues", digest)
        self.assertIn("$8.25M", digest)
        self.assertIn("5.99x", digest)

    def test_venture_financing_digest_preserves_vc_issue_model(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="emerging-companies-venture-capital/analyze-counterparty-markup-of-bridge-loan-agreement",
            question="Review the investor markup of our convertible note purchase agreement against the board-approved term sheet.",
            answer_schema={"deliverables": ["bridge-markup-analysis-memo.docx"]},
            metadata={"practice_area": "emerging-companies-venture-capital"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "board-approved-term-sheet.docx", "extension": ".docx"},
                {"doc_id": "doc_2", "filename": "cfv-markup-npa.docx", "extension": ".docx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "CFV commits $2,500,000 of a $4,500,000 bridge. Board guardrails say MFN must remain elective, "
                        "observer cannot attend executive sessions or committees, and side letters cannot conflict with the NPA."
                    ),
                },
                {
                    "doc_id": "doc_2",
                    "chunk_id": "c2",
                    "index": 1,
                    "text": (
                        "The markup adds a MAC definition covering prospects, a side letter that controls over inconsistent terms, "
                        "automatic MFN, $150,000 expenditure consent, $100,000 cross-default, committee attendance, and executive sessions."
                    ),
                },
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("High-Priority Venture Financing Issue Matrix", digest)
        self.assertIn("MAC Event of Default includes prospects", digest)
        self.assertIn("55.6%", digest)
        self.assertIn("Frankenstein notes", digest)
        self.assertIn("24.2%", digest)
        self.assertNotIn("Fund economics", digest)
        prompt = build_synthesis_prompt(state)
        self.assertIn("For emerging-company, venture-financing", prompt)
        self.assertIn("High-Priority Venture Financing Issue Matrix", prompt)

    def test_venture_charter_digest_requires_actual_drafting_terms(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="emerging-companies-venture-capital/draft-amended-and-restated-certificate-of-incorporation",
            question="Draft a second amended and restated certificate of incorporation for a Series B preferred stock financing.",
            answer_schema={"deliverables": ["second-amended-restated-coi.docx"]},
            metadata={"practice_area": "emerging-companies-venture-capital"},
        )
        state = RunState(
            task=task,
            config=load_config(),
            documents=[
                {"doc_id": "doc_1", "filename": "series-b-term-sheet.docx", "extension": ".docx"},
                {"doc_id": "doc_2", "filename": "second-amended-certificate.docx", "extension": ".docx"},
            ],
            chunks=[
                {
                    "doc_id": "doc_1",
                    "chunk_id": "c1",
                    "index": 0,
                    "text": (
                        "Series B has $4.00 original issue price, $75,000,000 Qualified IPO gross proceeds, "
                        "broad-based weighted average anti-dilution, and non-cumulative dividends."
                    ),
                }
            ],
        )
        digest = build_task_family_digest(state)
        self.assertIn("Certificate Drafting Requirements / Remediation Checklist", digest)
        self.assertIn("$5,000,000", digest)
        self.assertIn("31,250,000 total authorized shares", digest)
        self.assertIn("Do not increase total authorized shares for shadow series", digest)
        self.assertIn("CP2 = CP1 x (A + B) / (A + C)", digest)
        self.assertIn("$0.32/share/year", digest)
        self.assertIn("Series B receives $4.00/share first", digest)

    def test_antitrust_digest_routes_hsr_and_preserves_market_hot_doc_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="antitrust-competition/analyze-antitrust-hsr-strategy",
            question="Prepare an antitrust risk assessment and HSR filing strategy memo.",
            answer_schema={"deliverables": ["antitrust-risk-assessment-memo.docx", "hsr-filing-strategy-memo.docx"]},
            metadata={"practice_area": "antitrust-competition"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        contract = build_deliverable_contract(state)
        digest = build_task_family_digest(state)
        self.assertEqual(contract["task_family"], "antitrust_competition_review")
        self.assertIn(
            "Near-top antitrust / competition required findings",
            contract["deliverables"][0]["required_sections"],
        )
        self.assertIn("Market Definition / HHI / Share Matrix", digest)
        self.assertIn("Score-Critical Antitrust Preservation Checklist", digest)
        self.assertIn("post-merger HHI greater than 1,800 and delta HHI greater than 100", digest)
        self.assertIn("Greenville-Spartanburg", digest)
        self.assertIn("approximately 3,338", digest)
        self.assertIn("Initial HSR waiting period is 30 days", digest)
        self.assertIn("47 miles apart", digest)
        self.assertIn("third-largest U.S. distributor", digest)
        self.assertIn("specialty gases", digest)
        self.assertIn("eliminates pricing pressure", digest)
        self.assertIn("18-22 million dollars", digest)
        self.assertIn("47 of 156 bids", digest)
        self.assertIn("119.5 million dollars", digest)
        self.assertIn("TerraGas Industries", digest)
        self.assertIn("24.25 million dollars", digest)
        prompt = build_synthesis_prompt(state)
        self.assertIn("For antitrust, HSR, merger-risk", prompt)
        self.assertIn("Hot-Document and Bad-Fact Inventory", prompt)

    def test_anemic_synthesis_guard_preserves_worker_packet_for_dense_tasks(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="antitrust-competition/compare-corporate-antitrust-compliance-program-against-doj-and-ftc-guidelines",
            question="Compare antitrust compliance program against DOJ and FTC guidelines.",
            answer_schema={"deliverables": ["antitrust-compliance-gap-analysis.docx"], "criteria_count": 63},
            metadata={"practice_area": "antitrust-competition"},
        )
        state = RunState(task=task, config=load_config(), documents=[])
        state.final_packet = {
            "cheap_worker_summary": "CCO reporting structure lacks board independence; response time was 34 business days.",
            "deliverable_contract": build_deliverable_contract(state),
        }
        short_draft = "Brief memo: improve the compliance program."
        self.assertTrue(is_anemic_synthesis_answer(state, short_draft))
        fallback = build_anemic_synthesis_fallback_answer(state, short_draft)
        self.assertIn("CCO reporting structure lacks board independence", fallback)
        self.assertIn("34 business days", fallback)

    def test_antitrust_digest_routes_protective_order_compliance_and_expert_rows(self) -> None:
        protective_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="antitrust-competition/analyze-counterparty-markup-of-protective-order",
            question="Analyze counterparty markup of protective order.",
            answer_schema={"deliverables": ["protective-order-markup-memo.docx"]},
            metadata={"practice_area": "antitrust-competition"},
        )
        protective_digest = build_task_family_digest(RunState(task=protective_task, config=load_config(), documents=[]))
        self.assertIn("Protective-Order Clause-Delta Matrix", protective_digest)
        self.assertIn("government agency personnel", protective_digest)
        self.assertIn("44 U.S.C. 3301", protective_digest)
        self.assertIn("OptiPrice patent portfolio", protective_digest)
        self.assertIn("DOC_005", protective_digest)
        self.assertIn("DOC_003", protective_digest)
        self.assertIn("Kamakana", protective_digest)

        compliance_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="antitrust-competition/compare-corporate-antitrust-compliance-program-against-doj-and-ftc-guidelines",
            question="Compare antitrust compliance program against DOJ and FTC guidelines.",
            answer_schema={"deliverables": ["antitrust-compliance-gap-analysis.docx"]},
            metadata={"practice_area": "antitrust-competition"},
        )
        compliance_digest = build_task_family_digest(RunState(task=compliance_task, config=load_config(), documents=[]))
        self.assertIn("Compliance Program Gap Matrix", compliance_digest)
        self.assertIn("Design, Implementation, and Effectiveness", compliance_digest)
        self.assertIn("Frank J. Bellingham", compliance_digest)
        self.assertIn("Kenji Watanabe", compliance_digest)
        self.assertIn("34 business days", compliance_digest)
        self.assertIn("11 countries", compliance_digest)
        self.assertIn("March 15, 2018", compliance_digest)
        self.assertIn("4.7 billion dollars", compliance_digest)
        self.assertIn("since 2020", compliance_digest)

        expert_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="antitrust-competition/compare-expert-market-share-estimates-against-agency-data",
            question="Compare expert market share estimates against agency data.",
            answer_schema={"deliverables": ["market-share-reconciliation-memo.docx"]},
            metadata={"practice_area": "antitrust-competition"},
        )
        expert_digest = build_task_family_digest(RunState(task=expert_task, config=load_config(), documents=[]))
        self.assertIn("Expert / Agency Data Reconciliation Matrix", expert_digest)
        self.assertIn("15.0% to 17.0%", expert_digest)
        self.assertIn("Brentwood Foods tolling arrangement", expert_digest)
        self.assertIn("fringe-firm assumption", expert_digest)
        self.assertIn("$4.8B versus FTC $3.6B", expert_digest)
        self.assertIn("identical at $14.2B", expert_digest)
        self.assertIn("Freedonia Group", expert_digest)
        self.assertIn("Census / NAICS", expert_digest)
        self.assertIn("1,396", expert_digest)

    def test_antitrust_digest_routes_iss_transaction_rows(self) -> None:
        task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="antitrust-competition/analyze-iss-antitrust-transaction-structure",
            question="Analyze ISS antitrust transaction structure.",
            answer_schema={"deliverables": ["antitrust-transaction-structure-memo.docx"]},
            metadata={"practice_area": "antitrust-competition"},
        )
        digest = build_task_family_digest(RunState(task=task, config=load_config(), documents=[]))
        self.assertIn("ISS / Transaction Structure Antitrust Rows", digest)
        self.assertIn("60 million dollar divestiture cap", digest)
        self.assertIn("82 million dollars", digest)
        self.assertIn("256.2 million dollars", digest)
        self.assertIn("215.2 million dollars", digest)
        self.assertIn("coordinated effects", digest)
        self.assertIn("14 of Pinnacle's top 20 customers", digest)
        self.assertIn("Pinnacle as Aldersgate's primary competitive constraint", digest)
        self.assertIn("Janet Holbrook", digest)
        self.assertIn("$14.2M pricing-synergy breakdown", digest)
        self.assertIn("$23.2M", digest)
        self.assertIn("Clayton Act Section 7", digest)

    def test_litigation_digest_routes_motion_discovery_hold_and_invoice_rows(self) -> None:
        motion_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="litigation-dispute-resolution/analyze-counterparty-motion-to-dismiss",
            question="Review the attached motion to dismiss and produce an opposition issues memo.",
            answer_schema={"deliverables": ["motion-to-dismiss-issue-memo.docx"]},
            metadata={"practice_area": "litigation-dispute-resolution"},
        )
        motion_state = RunState(task=motion_task, config=load_config(), documents=[])
        motion_digest = build_task_family_digest(motion_state)
        self.assertIn("Motion Procedure / Authority Matrix", build_synthesis_prompt(motion_state))
        self.assertIn("Atlantic Marine", motion_digest)
        self.assertIn("Section 1404(a)", motion_digest)
        self.assertIn("100 Techwood Drive NW", motion_digest)
        self.assertIn("14 employees", motion_digest)
        self.assertIn("Rule 12(b)(6)", motion_digest)
        self.assertIn("Statement of Work Capabilities", motion_digest)
        self.assertIn("GUDTPA", motion_digest)
        self.assertIn("Anil Venkatesh", motion_digest)

        discovery_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="litigation-dispute-resolution/analyze-counterparty-requests-for-production-for-objectionable-and-overbroad-discovery-demands",
            question="Analyze requests for production for objectionable and overbroad discovery demands.",
            answer_schema={"deliverables": ["discovery-objections-memo.docx"]},
            metadata={"practice_area": "litigation-dispute-resolution"},
        )
        discovery_digest = build_task_family_digest(RunState(task=discovery_task, config=load_config(), documents=[]))
        self.assertIn("RFP No. 36", discovery_digest)
        self.assertIn("RFP No. 41", discovery_digest)
        self.assertIn("Fed. R. Civ. P. 33(a)(2)", discovery_digest)
        self.assertIn("FRE 502(d)", discovery_digest)
        self.assertIn("September 11, 2024", discovery_digest)
        self.assertIn("Section 9.1", discovery_digest)

        hold_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="litigation-dispute-resolution/assess-litigation-hold-scope-for-custodian-identification",
            question="Assess litigation hold scope for custodian identification.",
            answer_schema={"deliverables": ["litigation-hold-scope-memo.docx"]},
            metadata={"practice_area": "litigation-dispute-resolution"},
        )
        hold_digest = build_task_family_digest(RunState(task=hold_task, config=load_config(), documents=[]))
        self.assertIn("Renata Sokolova", hold_digest)
        self.assertIn("Tomás Herrera", hold_digest)
        self.assertIn("Monica Tran-Nguyen", hold_digest)
        self.assertIn("SOX Section 806", hold_digest)
        self.assertIn("Graham Ellicott", hold_digest)
        self.assertIn("Diana Muñoz", hold_digest)
        self.assertIn("Frank Jessup", hold_digest)

        invoice_task = BenchmarkTask(
            benchmark="harvey_lab_sample",
            task_id="litigation-dispute-resolution/assess-reasonableness-of-staffing-levels-on-litigation-invoice",
            question="Assess reasonableness of staffing levels on litigation invoice.",
            answer_schema={"deliverables": ["invoice-review-memo.docx"]},
            metadata={"practice_area": "litigation-dispute-resolution"},
        )
        invoice_digest = build_task_family_digest(RunState(task=invoice_task, config=load_config(), documents=[]))
        self.assertIn("$10,312.50", invoice_digest)
        self.assertIn("Section 5.4", invoice_digest)
        self.assertIn("38.5 hours", invoice_digest)
        self.assertIn("$95,000-$165,000", invoice_digest)
        self.assertIn("$2,125.00", invoice_digest)
        self.assertIn("$3,487.50", invoice_digest)
        self.assertIn("$33,600.63", invoice_digest)


if __name__ == "__main__":
    unittest.main()
